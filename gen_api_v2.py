"""
SmithLife API 명세서 v2.0 - 누락 보완 완전판
Part 1: Sheet1(총괄), Sheet3(에러코드), Sheet4(공통규격), Sheet5(DB매핑)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 스타일 ──
DARK = "2C2C2C"; WHITE = "FFFFFF"; GOLD = "D4A04A"
LIGHT_GRAY = "F5F5F5"; BLUE = "2980B9"; GREEN = "27AE60"; RED = "E74C3C"; ORANGE = "E67E22"

hf = Font(name="맑은 고딕", bold=True, size=11, color=WHITE)
hfill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
bf = Font(name="맑은 고딕", size=10)
wa = Alignment(wrap_text=True, vertical="top", horizontal="left")
ca = Alignment(wrap_text=True, vertical="center", horizontal="center")
tb = Border(left=Side("thin","CCCCCC"),right=Side("thin","CCCCCC"),top=Side("thin","CCCCCC"),bottom=Side("thin","CCCCCC"))
af = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
MF = {
    "GET": Font(name="맑은 고딕",size=10,bold=True,color=BLUE),
    "POST": Font(name="맑은 고딕",size=10,bold=True,color=GREEN),
    "PUT": Font(name="맑은 고딕",size=10,bold=True,color=ORANGE),
    "PATCH": Font(name="맑은 고딕",size=10,bold=True,color=ORANGE),
    "DELETE": Font(name="맑은 고딕",size=10,bold=True,color=RED),
}

def style_hdr(ws, hdrs, ws_list):
    for c,(h,w) in enumerate(zip(hdrs,ws_list),1):
        cell=ws.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.alignment=ca; cell.border=tb
        ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="A2"

def write_rows(ws, data, method_col=None):
    for r,row in enumerate(data,2):
        for c,val in enumerate(row,1):
            cell=ws.cell(row=r,column=c,value=val); cell.font=bf; cell.alignment=wa; cell.border=tb
            if r%2==0: cell.fill=af
        if method_col:
            mc=ws.cell(row=r,column=method_col)
            m=str(mc.value).upper() if mc.value else ""
            if m in MF: mc.font=MF[m]; mc.alignment=ca

# ============================================================
# Sheet 1: API 총괄 목록 (90개 - 누락 11개 추가)
# ============================================================
ws1 = wb.active; ws1.title = "API 총괄 목록"
H1=["No","대분류","중분류","API ID","HTTP Method","엔드포인트 (URI)","기능명","설명","인증","관련 SRS ID"]
W1=[5,9,11,14,9,40,16,52,8,22]

api_list = [
    # ── 인증 (AUTH) 10개 ──
    [1,"인증","로그인","AUTH-API-001","POST","/api/v1/auth/login","로그인","이메일/비밀번호로 로그인, JWT(Access+Refresh) 토큰 발급","X","AUTH-001~006"],
    [2,"인증","세션","AUTH-API-002","POST","/api/v1/auth/token/refresh","토큰 갱신","Refresh Token으로 새 Access Token 발급","O(Refresh)","SPL-003, NFR-007"],
    [3,"인증","세션","AUTH-API-003","POST","/api/v1/auth/token/verify","토큰 유효성 검증","앱 시작 시 저장된 Access Token 유효 여부 확인","O","SPL-003, SPL-005"],
    [4,"인증","로그아웃","AUTH-API-004","POST","/api/v1/auth/logout","로그아웃","서버 측 Refresh Token 무효화(블랙리스트) 처리","O","SET-007"],
    [5,"인증","회원가입","AUTH-API-005","POST","/api/v1/auth/signup","회원가입","이름/전화번호/이메일/비밀번호로 신규 회원 등록","X","AUTH-009~016"],
    [6,"인증","회원가입","AUTH-API-006","GET","/api/v1/auth/check-email?email={email}","이메일 중복 확인","가입 시 이메일 중복 여부 실시간 체크","X","AUTH-012, AUTH-015"],
    [7,"인증","회원가입","AUTH-API-007","GET","/api/v1/auth/check-phone?phone={phone}","전화번호 중복 확인","가입 시 전화번호 중복 여부 실시간 체크","X","AUTH-011, AUTH-015"],
    [8,"인증","계정 찾기","AUTH-API-008","POST","/api/v1/auth/find-email","아이디(이메일) 찾기","전화번호로 등록된 이메일 조회 (마스킹 처리)","X","AUTH-018"],
    [9,"인증","계정 찾기","AUTH-API-009","POST","/api/v1/auth/reset-password/request","비밀번호 재설정 요청","등록된 이메일로 재설정 코드 발송 (5분 유효)","X","AUTH-019"],
    [10,"인증","계정 찾기","AUTH-API-010","POST","/api/v1/auth/reset-password/confirm","비밀번호 재설정 확인","인증 코드 확인 후 새 비밀번호 설정","X","AUTH-019"],

    # ── 홈 (HOME) 2개 ──
    [11,"홈","오늘의 요약","HOME-API-001","GET","/api/v1/home/summary","홈 요약 조회","다음 예약, 혼잡도, 회원권 D-day 등 홈 화면 통합 데이터","O","HOME-003~006,010"],
    [12,"홈","공지사항","HOME-API-002","GET","/api/v1/home/announcements?limit={n}","홈 공지 미리보기","홈 가로 스크롤용 최신 공지 N건 조회","O","HOME-007,008"],

    # ── 사용자 (USER) 5개 ──
    [13,"사용자","프로필","USER-API-001","GET","/api/v1/users/me","내 정보 조회","프로필+회원권+참석율+오늘 운동량 통합 조회","O","MY-001~003"],
    [14,"사용자","프로필","USER-API-002","PUT","/api/v1/users/me","내 정보 수정","이름, 전화번호, 프로필 이미지 URL 등 수정","O","MY-007, SET-002"],
    [15,"사용자","프로필","USER-API-003","POST","/api/v1/users/me/profile-image","프로필 이미지 업로드","프로필 사진 파일 업로드 (multipart/form-data)","O","MY-007"],
    [16,"사용자","계정","USER-API-004","PUT","/api/v1/users/me/password","비밀번호 변경","현재 비밀번호 확인 후 새 비밀번호 설정","O","SET-006"],
    [17,"사용자","계정","USER-API-005","DELETE","/api/v1/users/me","회원 탈퇴","비밀번호 재확인 후 계정 비활성화(soft delete)","O","SET-008"],

    # ── 회원권 (MEMBERSHIP) 3개 ──
    [18,"회원권","조회","MBSP-API-001","GET","/api/v1/memberships/me","내 회원권 조회","현재 활성 회원권 (종류/시작일/만료일/D-day) 조회","O","HOME-006, MY-001"],
    [19,"회원권","관리","MBSP-API-002","POST","/api/v1/memberships","회원권 등록/연장","신규 회원권 등록 또는 기존 회원권 연장","O","NOTI-004"],
    [20,"회원권","관리","MBSP-API-003","PATCH","/api/v1/memberships/{membershipId}/pause","회원권 일시정지","회원권 PAUSED 상태 변경 (최대 30일)","O","FAQ"],

    # ── 출석/QR (ATTENDANCE) 5개 ──
    [21,"출석","QR 체크인","ATT-API-001","POST","/api/v1/attendance/qr-token","QR 토큰 생성","사용자 고유 1회성 QR 코드 토큰 생성 (60초 유효)","O","QR-003,004, NFR-009"],
    [22,"출석","QR 체크인","ATT-API-002","POST","/api/v1/attendance/check-in","QR 체크인 처리","QR 스캔 후 출석 체크인 기록 생성 (키오스크/직원용)","O(Admin)","QR-005"],
    [23,"출석","QR 체크인","ATT-API-003","PATCH","/api/v1/attendance/{attendanceId}/check-out","체크아웃 처리","퇴장 시 체크아웃 시간 기록","O","QR-005"],
    [24,"출석","조회","ATT-API-004","GET","/api/v1/attendance/me?month={yyyy-MM}","내 출석 기록 조회","월별 출석 기록 리스트 조회 (참석율 계산용)","O","MY-002"],
    [25,"출석","조회","ATT-API-005","GET","/api/v1/attendance/me/status","현재 출석 상태","현재 체크인 여부 및 진행 중 출석 정보 조회","O","QR-005"],

    # ── 시설 (FACILITY) 2개 ──
    [26,"시설","조회","FAC-API-001","GET","/api/v1/facilities","시설 목록 조회","전체 시설(메인 피트니스 존, 스튜디오 A/B 등) 목록","O","HOME-004, RESV-001"],
    [27,"시설","혼잡도","FAC-API-002","GET","/api/v1/facilities/{facilityId}/congestion","시설 혼잡도 조회","특정 시설 현재 혼잡도(원활/보통/혼잡) 실시간 조회","O","HOME-005, RESV-003"],

    # ── 시간 슬롯 (TIME SLOT) 1개 ──
    [28,"예약","시간표","SLOT-API-001","GET","/api/v1/facilities/{facilityId}/slots?date={yyyy-MM-dd}","시간 슬롯 조회","특정 시설/날짜의 시간 슬롯 + 혼잡도 + 내 예약 여부","O","RESV-002,003,006"],

    # ── 예약 (RESERVATION) 6개 ──
    [29,"예약","예약","RESV-API-001","POST","/api/v1/reservations","예약 생성","선택한 시간 슬롯(복수 가능)으로 예약 생성, 예약번호 발급","O","RESV-009~012"],
    [30,"예약","예약","RESV-API-002","GET","/api/v1/reservations/me?date={yyyy-MM-dd}&facilityId={id}","내 예약 목록 조회","날짜/시설 필터 기준 내 예약 리스트 (취소 모달용)","O","RESV-005,014, HOME-004"],
    [31,"예약","예약","RESV-API-003","GET","/api/v1/reservations/me/upcoming","다음 예약 조회","가장 가까운 미래 예약 1건 (홈 카드용)","O","HOME-004"],
    [32,"예약","취소","RESV-API-004","PATCH","/api/v1/reservations/{reservationId}/cancel","예약 취소","예약 CANCELLED 변경, 슬롯 인원 차감","O","RESV-015~017"],
    [33,"예약","상세","RESV-API-005","GET","/api/v1/reservations/{reservationId}","예약 상세 조회","예약번호/시설/날짜/시간/상태 등 상세 정보","O","NOTI-006"],
    [34,"예약","캘린더","RESV-API-006","GET","/api/v1/reservations/me/calendar?month={yyyy-MM}","월별 예약 캘린더","날짜별 예약 유무 마커 데이터 (캘린더 dot 표시용)","O","RESV-006"],

    # ── 이용내역 (USAGE) 2개 ──
    [35,"이용내역","조회","USAGE-API-001","GET","/api/v1/usage-history/me?page={p}&size={s}&status={st}&from={date}&to={date}","이용내역 목록 조회","최신순 페이지네이션, 상태/기간 필터 지원 (무한스크롤)","O","USAGE-001~006, HOME-009"],
    [36,"이용내역","상세","USAGE-API-002","GET","/api/v1/usage-history/{usageId}","이용내역 상세","개별 이용내역 상세(시설명/시간/예약번호/상태)","O","USAGE-002"],

    # ── 운동 종목 (EXERCISE) 3개 ──
    [37,"운동","종목","EXER-API-001","GET","/api/v1/exercises?bodyPart={part}&equipment={eq}&keyword={kw}","운동 종목 검색","부위/장비/키워드 필터로 운동 종목 목록 검색","O","WORK-009, RTNC-005"],
    [38,"운동","종목","EXER-API-002","GET","/api/v1/exercises/{exerciseId}","운동 종목 상세","운동명/부위/장비/이미지/설명 등 상세 정보","O","EXER-001"],
    [39,"운동","종목","EXER-API-003","GET","/api/v1/exercises/{exerciseId}/history?limit={n}","운동 이전 기록 조회","해당 운동의 최근 수행 이력(날짜/무게/횟수) N건","O","EXER-006"],

    # ── 루틴 (ROUTINE) 7개 ──
    [40,"운동","추천 루틴","RTN-API-001","GET","/api/v1/routines/recommended","추천 루틴 목록","시스템 추천 루틴 목록 (2열 그리드용)","O","RTN-001~005, WORK-004~005"],
    [41,"운동","루틴 검색","RTN-API-002","GET","/api/v1/routines?keyword={kw}&goal={goal}&difficulty={diff}","루틴 검색","키워드/목표/난이도 필터 루틴 검색 (공개 루틴 포함)","O","RTN-002"],
    [42,"운동","루틴 상세","RTN-API-003","GET","/api/v1/routines/{routineId}","루틴 상세 조회","루틴 기본정보+운동 구성 리스트+찜 여부 통합","O","RTND-001~005"],
    [43,"운동","루틴 생성","RTN-API-004","POST","/api/v1/routines","나만의 루틴 생성","루틴명/목표/운동목록(순서포함)/공유여부 저장","O","RTNC-001~008"],
    [44,"운동","루틴 관리","RTN-API-005","PUT","/api/v1/routines/{routineId}","루틴 수정","이름/목표/운동구성(순서변경)/공유 설정 수정","O","RTNC-006"],
    [45,"운동","루틴 관리","RTN-API-006","DELETE","/api/v1/routines/{routineId}","루틴 삭제","내가 만든 루틴 삭제","O","RTNC-001"],
    [46,"운동","나의 루틴","RTN-API-007","GET","/api/v1/routines/me","내 루틴 목록","내가 생성한 루틴 목록 조회","O","RTN-006"],

    # ── 루틴 찜 (FAVORITE) 3개 ──
    [47,"운동","찜","FAV-API-001","POST","/api/v1/routines/{routineId}/favorite","루틴 찜 추가","루틴을 즐겨찾기에 추가 (하트 채우기)","O","RTND-006"],
    [48,"운동","찜","FAV-API-002","DELETE","/api/v1/routines/{routineId}/favorite","루틴 찜 해제","루틴을 즐겨찾기에서 제거 (하트 비우기)","O","RTND-006"],
    [49,"운동","찜","FAV-API-003","GET","/api/v1/routines/favorites","찜한 루틴 목록","내가 찜한 루틴 전체 목록 조회","O","RTND-006"],

    # ── 운동 세션 (WORKOUT SESSION) 5개 ──
    [50,"운동","세션","SESS-API-001","POST","/api/v1/workout-sessions","운동 세션 시작","새 운동 세션 생성 (루틴 기반 or 빈 세션)","O","RTND-007, WORK-011"],
    [51,"운동","세션","SESS-API-002","GET","/api/v1/workout-sessions/active","활성 세션 조회","현재 ACTIVE/PAUSED 세션 + 운동 리스트 조회","O","WORK-001~003,006"],
    [52,"운동","세션","SESS-API-003","PATCH","/api/v1/workout-sessions/{sessionId}/end","세션 종료","COMPLETED 변경, 총시간/볼륨/칼로리 최종 계산","O","WORK-010"],
    [53,"운동","세션","SESS-API-004","PATCH","/api/v1/workout-sessions/{sessionId}/pause","세션 일시정지","세션 PAUSED 상태 변경","O","WORK-001"],
    [54,"운동","세션","SESS-API-005","PATCH","/api/v1/workout-sessions/{sessionId}/resume","세션 재개","세션 ACTIVE 상태 재변경","O","WORK-001"],

    # ── 세션 내 운동 (SESSION EXERCISE) 5개 ──
    [55,"운동","세션 운동","SE-API-001","POST","/api/v1/workout-sessions/{sessionId}/exercises","세션에 운동 추가","진행 중 세션에 새 운동 추가","O","WORK-009"],
    [56,"운동","세션 운동","SE-API-002","GET","/api/v1/workout-sessions/{sessionId}/exercises","세션 운동 목록","세션 내 운동 리스트 (완료/진행중/대기 상태)","O","WORK-006~008"],
    [57,"운동","세션 운동","SE-API-003","PATCH","/api/v1/session-exercises/{seId}/complete","세션 운동 완료","특정 운동 COMPLETED 상태 변경","O","EXER-008"],
    [58,"운동","세션 운동","SE-API-004","PATCH","/api/v1/session-exercises/{seId}/start","세션 운동 시작","IN_PROGRESS(UP NEXT→진행중) 변경","O","WORK-008"],
    [59,"운동","세션 운동","SE-API-005","DELETE","/api/v1/session-exercises/{seId}","세션 운동 삭제","세션에서 특정 운동 제거","O","WORK-009"],

    # ── 세트 기록 (EXERCISE SET) 5개 ──
    [60,"운동","세트","SET-API-001","POST","/api/v1/session-exercises/{seId}/sets","세트 추가","운동에 새 세트 행 추가 (+ 세트 추가하기)","O","EXER-005"],
    [61,"운동","세트","SET-API-002","GET","/api/v1/session-exercises/{seId}/sets","세트 목록 조회","전체 세트(번호/KG/회/완료여부) 리스트","O","EXER-002"],
    [62,"운동","세트","SET-API-003","PATCH","/api/v1/exercise-sets/{setId}","세트 정보 수정","KG/횟수/휴식시간 등 세트 데이터 수정","O","EXER-003"],
    [63,"운동","세트","SET-API-004","PATCH","/api/v1/exercise-sets/{setId}/complete","세트 완료 체크","세트 완료(초록체크)+완료시각+휴식타이머 트리거","O","EXER-004,009"],
    [64,"운동","세트","SET-API-005","DELETE","/api/v1/exercise-sets/{setId}","세트 삭제","특정 세트 행 삭제","O","EXER-005"],

    # ── 개인 기록 (PERSONAL RECORD) 2개 ──
    [65,"운동","개인기록","PR-API-001","GET","/api/v1/personal-records?exerciseId={id}","운동별 최고 기록","특정 운동의 1RM/최대볼륨/최대횟수 조회","O","EXER-006"],
    [66,"운동","개인기록","PR-API-002","GET","/api/v1/personal-records/me","전체 개인 기록","내 모든 운동의 개인 최고 기록 리스트","O","EXER-006"],

    # ── 리포트 (WORKOUT REPORT) 4개 ──
    [67,"리포트","요약","RPT-API-001","GET","/api/v1/workout-reports/{sessionId}","운동 리포트 조회","세션 완료 후 리포트(시간/볼륨/칼로리/운동요약)","O","RPT-001~004,006"],
    [68,"리포트","차트","RPT-API-002","GET","/api/v1/workout-reports/weekly?date={yyyy-MM-dd}","주간 운동량 차트","월~일 요일별 운동량 + 지난주 대비 변화율(%)","O","RPT-003"],
    [69,"리포트","공유","RPT-API-003","POST","/api/v1/workout-reports/{reportId}/share","리포트 공유","리포트 이미지 캡처/생성 → 공유 URL 반환","O","RPT-005"],
    [70,"리포트","오늘","RPT-API-004","GET","/api/v1/workout-reports/today","오늘의 운동량","오늘 총 볼륨(kg) 등 내정보 카드용 데이터","O","MY-003"],

    # ── 설정 (SETTINGS) 2개 ──
    [71,"설정","사용자 설정","SETT-API-001","GET","/api/v1/settings","설정 조회","알림/다크모드/언어 등 현재 설정값 조회","O","SET-003~005"],
    [72,"설정","사용자 설정","SETT-API-002","PATCH","/api/v1/settings","설정 변경","알림ON/OFF, 다크모드(SYSTEM/ON/OFF), 언어 변경","O","SET-003~005, SET-012"],

    # ── 알림 (NOTIFICATION) 4개 ──
    [73,"알림","개인 알림","NOTI-API-001","GET","/api/v1/notifications?page={p}&size={s}&type={type}","알림 목록 조회","유형 필터 지원 페이지네이션 알림 목록 (읽음/미읽음 구분)","O","NOTI-001~005"],
    [74,"알림","개인 알림","NOTI-API-002","PATCH","/api/v1/notifications/{notificationId}/read","알림 읽음 처리","개별 알림 읽음 상태 변경","O","NOTI-005"],
    [75,"알림","개인 알림","NOTI-API-003","PATCH","/api/v1/notifications/read-all","전체 알림 읽음","모든 미읽음 알림 일괄 읽음 처리","O","NOTI-002"],
    [76,"알림","개인 알림","NOTI-API-004","GET","/api/v1/notifications/unread-count","미읽음 알림 수","벨 아이콘 배지 숫자용 미읽음 개수","O","HOME-002"],

    # ── 공지사항 (ANNOUNCEMENT) 2개 ──
    [77,"공지사항","목록","ANN-API-001","GET","/api/v1/announcements?page={p}&size={s}&tag={tag}","공지사항 목록","태그(NOTICE/EVENT) 필터 페이지네이션 공지 리스트","O","NOTI-007, SET-009"],
    [78,"공지사항","상세","ANN-API-002","GET","/api/v1/announcements/{announcementId}","공지사항 상세","공지 전체 내용(제목/내용/이미지/게시일) 조회","O","NOTI-008, HOME-008"],

    # ── 1:1 문의 (INQUIRY) 4개 ──
    [79,"문의","1:1 문의","INQ-API-001","GET","/api/v1/inquiries/me?page={p}&size={s}&period={m}&status={st}","내 문의 목록","기간/상태 필터 문의 내역 카드 리스트","O","INQ-004~006"],
    [80,"문의","1:1 문의","INQ-API-002","POST","/api/v1/inquiries","새 문의 작성","제목/내용/카테고리/이미지첨부로 1:1 문의 접수","O","INQ-007~009"],
    [81,"문의","1:1 문의","INQ-API-003","GET","/api/v1/inquiries/{inquiryId}","문의 상세 조회","문의 내용 + 답변 내용(있으면) 통합 조회","O","INQ-006"],
    [82,"문의","이미지","INQ-API-004","POST","/api/v1/inquiries/upload-image","문의 이미지 업로드","문의 첨부 이미지 파일 업로드 (multipart)","O","INQ-008"],

    # ── FAQ 1개 ──
    [83,"문의","FAQ","FAQ-API-001","GET","/api/v1/faqs?category={cat}","FAQ 목록 조회","카테고리 필터 자주 묻는 질문 전체 목록 (아코디언)","O","INQ-002"],

    # ── 챗봇 (CHATBOT) 2개 [신규] ──
    [84,"문의","챗봇","CHAT-API-001","POST","/api/v1/chatbot/sessions","챗봇 세션 생성","새 챗봇 상담 세션 시작","O","INQ-003"],
    [85,"문의","챗봇","CHAT-API-002","POST","/api/v1/chatbot/sessions/{chatSessionId}/messages","챗봇 메시지 전송","사용자 메시지 전송 → AI/규칙기반 응답 반환","O","INQ-003"],

    # ── 디바이스/푸시 (DEVICE) 2개 [신규] ──
    [86,"알림","푸시","DEV-API-001","POST","/api/v1/devices/fcm-token","FCM 토큰 등록","디바이스 FCM/APNs 푸시 토큰 등록/갱신","O","NOTI-009"],
    [87,"알림","푸시","DEV-API-002","DELETE","/api/v1/devices/fcm-token","FCM 토큰 삭제","로그아웃 시 디바이스 푸시 토큰 삭제","O","NOTI-009"],

    # ── 앱 공통 (APP) 3개 [신규] ──
    [88,"공통","앱","APP-API-001","GET","/api/v1/app/version-check?platform={ios|android}&version={ver}","앱 버전 체크","최소 지원 버전 비교, 강제 업데이트 여부 반환","X","SET-011, NFR-016"],
    [89,"공통","약관","APP-API-002","GET","/api/v1/app/terms?type={service|privacy}","이용약관/정책 조회","이용약관 또는 개인정보처리방침 HTML/Markdown 조회","X","SET-010"],
    [90,"공통","약관","APP-API-003","GET","/api/v1/app/terms/versions","약관 버전 목록","약관 변경 이력 목록 조회","X","SET-010"],
]

style_hdr(ws1,H1,W1)
write_rows(ws1,api_list,method_col=5)

# ============================================================
# Sheet 3: 공통 에러 코드
# ============================================================
ws3 = wb.create_sheet("공통 에러 코드")
H3=["HTTP 상태코드","에러 코드","메시지 (ko)","설명","대응 방법"]
W3=[16,20,32,42,35]

err = [
    ["200 OK","-","성공","요청 정상 처리","-"],
    ["201 Created","-","생성 완료","리소스 생성 성공 (회원가입, 예약 등)","-"],
    ["204 No Content","-","삭제 완료","리소스 삭제 성공 (찜 해제, 세트 삭제 등)","-"],
    ["400 Bad Request","INVALID_INPUT","입력값이 올바르지 않습니다.","필수 값 누락, 형식 오류, 유효성 검증 실패","입력 필드 유효성 재확인"],
    ["400 Bad Request","INVALID_FORMAT","형식이 올바르지 않습니다.","이메일/전화번호 등 형식 오류","정규식 기반 클라이언트 검증"],
    ["400 Bad Request","PASSWORD_TOO_SHORT","비밀번호는 6자 이상이어야 합니다.","비밀번호 최소 길이 미충족","6자 이상 재입력 안내"],
    ["400 Bad Request","INVALID_PASSWORD","현재 비밀번호가 올바르지 않습니다.","비밀번호 변경/탈퇴 시 현재 비밀번호 불일치","현재 비밀번호 재입력"],
    ["400 Bad Request","ALREADY_CANCELLED","이미 취소된 예약입니다.","취소된 예약 재취소 시도","예약 상태 재확인"],
    ["400 Bad Request","SESSION_ALREADY_ENDED","이미 종료된 세션입니다.","종료된 세션 수정 시도","새 세션 시작 안내"],
    ["401 Unauthorized","INVALID_CREDENTIALS","아이디 또는 비밀번호가 올바르지 않습니다.","로그인 실패","재입력 안내"],
    ["401 Unauthorized","TOKEN_EXPIRED","인증이 만료되었습니다.","Access Token 만료","Refresh Token으로 갱신"],
    ["401 Unauthorized","TOKEN_INVALID","유효하지 않은 인증입니다.","JWT 위변조/무효","로그인 화면 이동"],
    ["401 Unauthorized","REFRESH_TOKEN_EXPIRED","재인증이 필요합니다.","Refresh Token 만료","로그인 화면 이동"],
    ["401 Unauthorized","RESET_CODE_INVALID","인증 코드가 올바르지 않습니다.","비밀번호 재설정 코드 불일치/만료","코드 재확인 또는 재발송"],
    ["403 Forbidden","ACCESS_DENIED","접근 권한이 없습니다.","타인 리소스 접근 / 권한 부족","권한 확인"],
    ["403 Forbidden","MEMBERSHIP_EXPIRED","회원권이 만료되었습니다.","만료 상태에서 시설 이용 시도","회원권 연장 안내"],
    ["403 Forbidden","ACCOUNT_DEACTIVATED","비활성화된 계정입니다.","탈퇴/정지 계정 로그인 시도","고객센터 문의 안내"],
    ["404 Not Found","RESOURCE_NOT_FOUND","요청한 정보를 찾을 수 없습니다.","존재하지 않는 리소스 ID","ID 재확인"],
    ["404 Not Found","USER_NOT_FOUND","등록되지 않은 사용자입니다.","이메일/전화번호 미등록","회원가입 안내"],
    ["404 Not Found","NO_ACTIVE_SESSION","활성 운동 세션이 없습니다.","활성 세션 조회 시 세션 없음","새 세션 시작 안내"],
    ["409 Conflict","DUPLICATE_EMAIL","이미 가입된 이메일입니다.","회원가입 이메일 중복","로그인 화면 유도"],
    ["409 Conflict","DUPLICATE_PHONE","이미 등록된 전화번호입니다.","전화번호 중복","로그인 화면 유도"],
    ["409 Conflict","SLOT_FULL","해당 시간은 만석입니다.","예약 시 최대 인원 초과","다른 시간 안내"],
    ["409 Conflict","DUPLICATE_RESERVATION","이미 예약된 시간입니다.","동일 슬롯 중복 예약","기존 예약 확인"],
    ["409 Conflict","ACTIVE_SESSION_EXISTS","이미 진행 중인 세션이 있습니다.","활성 세션 존재 시 새 세션 생성","기존 세션 종료 후 재시도"],
    ["409 Conflict","ALREADY_FAVORITED","이미 찜한 루틴입니다.","중복 찜 시도","UI 상태 동기화"],
    ["413 Payload Too Large","FILE_TOO_LARGE","파일 크기가 초과되었습니다.","이미지 업로드 5MB 초과","파일 크기 줄여 재업로드"],
    ["415 Unsupported Media Type","INVALID_FILE_TYPE","지원하지 않는 파일 형식입니다.","허용: JPG/PNG/WEBP","이미지 형식 변경"],
    ["429 Too Many Requests","RATE_LIMITED","요청이 너무 많습니다.","API 호출 빈도 초과","잠시 대기 후 재시도"],
    ["500 Internal Server Error","SERVER_ERROR","서버 오류가 발생했습니다.","서버 내부 오류","재시도 + 관리자 문의"],
    ["503 Service Unavailable","SERVICE_UNAVAILABLE","서비스 점검 중입니다.","서버 점검/배포 중","잠시 후 재시도"],
]

style_hdr(ws3,H3,W3)
write_rows(ws3,err)

# ============================================================
# Sheet 4: 공통 규격
# ============================================================
ws4 = wb.create_sheet("공통 규격")
H4=["항목","내용","예시","비고"]
W4=[22,52,48,32]

common = [
    ["Base URL","https://api.smithlife.co.kr/api/v1","https://api.smithlife.co.kr/api/v1/auth/login","개발: http://localhost:8080/api/v1"],
    ["API 버전","URL Path 방식 (v1)","/api/v1/...","버전 업 시 /api/v2/ 신설"],
    ["인증 방식","Bearer Token (JWT)",'Authorization: Bearer eyJhbG...',"Access: 1시간, Refresh: 14일"],
    ["Content-Type","application/json (기본)\nmultipart/form-data (파일 업로드)",'Content-Type: application/json; charset=utf-8',"이미지 업로드만 multipart"],
    ["Accept-Language","다국어 지원 (ko, en, ja)","Accept-Language: ko","에러 메시지 다국어 대응"],
    ["날짜 형식","ISO 8601","2026-02-17T14:30:00","Date: yyyy-MM-dd\nDateTime: yyyy-MM-dd'T'HH:mm:ss"],
    ["시간 형식","HH:mm (24시간제)","09:00, 14:30","예약 시간표/슬롯에 사용"],
    ["페이지네이션","Offset 기반 (page/size)","?page=0&size=20","page: 0부터, size: 기본 20"],
    ["페이지 응답","content/totalElements/totalPages/hasNext",'{"content":[],"totalElements":0,"totalPages":0,"hasNext":false}',"Spring Page 호환"],
    ["정렬","기본 최신순 (createdAt DESC)","?sort=createdAt,desc","sort 파라미터 지원"],
    ["에러 응답","code/message/details",'{"code":"INVALID_INPUT","message":"...","details":{"field":"email"}}',"모든 4xx/5xx 동일 구조"],
    ["빈 배열","데이터 없을 시 빈 배열 (null X)",'{"content":[],"totalElements":0}',"Empty State UI 판단용"],
    ["Boolean","true/false (소문자)",'{"available":true}',"JSON boolean 표준"],
    ["Null 처리","값 없음 시 null (필드 생략 X)",'{"profileImageUrl":null}',"클라이언트 null 체크 필수"],
    ["파일 업로드","최대 5MB, JPG/PNG/WEBP","Content-Type: multipart/form-data","프로필/문의 이미지"],
    ["CORS","허용 Origin 화이트리스트","Access-Control-Allow-Origin: https://smithlife.co.kr","개발: localhost 허용"],
    ["Rate Limiting","IP/User 분당 60회 (기본)","X-RateLimit-Remaining: 55","비밀번호 재설정: 분당 3회"],
    ["HTTPS","TLS 1.2 이상 필수","https://...","NFR-006 준수"],
    ["요청 ID","모든 응답 헤더에 추적 ID 포함","X-Request-Id: uuid-v4","로그 추적/디버깅용"],
    ["타임아웃","클라이언트 30초 / 서버 60초","Connection Timeout: 30s","NFR-003 (95% 2초 이내)"],
]

style_hdr(ws4,H4,W4)
write_rows(ws4,common)

# ============================================================
# Sheet 5: DB-API 매핑
# ============================================================
ws5 = wb.create_sheet("DB-API 매핑")
H5=["No","DB 테이블","테이블 설명","관련 API ID 목록","CRUD","비고"]
W5=[5,20,22,50,10,28]

mapping = [
    [1,"user","회원","AUTH-API-001~010, USER-API-001~005","CRUD","가입/로그인/프로필/탈퇴"],
    [2,"membership","회원권","MBSP-API-001~003, HOME-API-001","CRU","등록/조회/일시정지"],
    [3,"user_settings","사용자 설정","SETT-API-001~002","RU","설정 조회/변경"],
    [4,"attendance","출석 기록","ATT-API-001~005","CRU","QR 체크인/아웃/조회"],
    [5,"facility","시설","FAC-API-001~002, SLOT-API-001","R","시설 목록/혼잡도"],
    [6,"time_slot","시간 슬롯","SLOT-API-001, RESV-API-001","RU","시간표 조회/예약 시 count"],
    [7,"reservation","예약","RESV-API-001~006, HOME-API-001","CRU","예약 생성/조회/취소/캘린더"],
    [8,"usage_history","이용내역","USAGE-API-001~002","CR","이용 기록 자동생성/조회"],
    [9,"exercise","운동 종목","EXER-API-001~003","R","검색/상세/이전기록"],
    [10,"routine","루틴","RTN-API-001~007","CRUD","CRUD + 추천"],
    [11,"routine_exercise","루틴-운동 구성","RTN-API-003~005","CRD","루틴 운동 구성 관리"],
    [12,"favorite_routine","루틴 찜","FAV-API-001~003","CRD","찜 추가/해제/목록"],
    [13,"workout_session","운동 세션","SESS-API-001~005","CRU","시작/종료/일시정지/재개"],
    [14,"session_exercise","세션-운동","SE-API-001~005","CRUD","세션 내 운동 관리"],
    [15,"exercise_set","세트 기록","SET-API-001~005","CRUD","추가/수정/완료/삭제"],
    [16,"personal_record","개인 최고 기록","PR-API-001~002, EXER-API-003","CR","세트 완료 시 자동갱신"],
    [17,"workout_report","운동 리포트","RPT-API-001~004","CR","세션 종료 시 자동생성"],
    [18,"notification","알림","NOTI-API-001~004","CRU","생성/조회/읽음처리"],
    [19,"announcement","공지사항","ANN-API-001~002, HOME-API-002","R","목록/상세"],
    [20,"inquiry","1:1 문의","INQ-API-001~004","CR","작성/목록/상세"],
    [21,"inquiry_reply","문의 답변","INQ-API-003","R","문의 상세 시 함께 조회"],
    [22,"faq","자주 묻는 질문","FAQ-API-001","R","FAQ 목록"],
    [23,"(신규) device_token","디바이스 토큰","DEV-API-001~002","CD","FCM 토큰 등록/삭제"],
    [24,"(신규) chatbot_session","챗봇 세션","CHAT-API-001~002","CR","챗봇 상담 세션/메시지"],
]

style_hdr(ws5,H5,W5)
write_rows(ws5,mapping)

# ── Sheet 2 자리 확보 (Part 2에서 채움) ──
ws2 = wb.create_sheet("API 상세 명세")
# 임시로 빈 시트 생성 - Part 2에서 완성
wb.move_sheet("API 상세 명세", offset=-3)  # 두 번째 위치로 이동

output_path = r"c:\SSAFY\personalProject\SmithLife2\SmithLife_API명세서_v2.0.xlsx"
wb.save(output_path)
print(f"Part 1 완료: {output_path}")
print(f"  Sheet 1: API 총괄 목록 - {len(api_list)}개 API")
print(f"  Sheet 3: 공통 에러 코드 - {len(err)}개")
print(f"  Sheet 4: 공통 규격 - {len(common)}개")
print(f"  Sheet 5: DB-API 매핑 - {len(mapping)}개")
print(f"  Sheet 2: 상세 명세 - Part 2에서 작성 예정")
