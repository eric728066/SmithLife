"""
SmithLife API 명세서 v2.0 - Part 2: 전체 90개 API 상세 명세
기존 v2.0.xlsx에 Sheet 2를 채워넣기
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook(r"c:\SSAFY\personalProject\SmithLife2\SmithLife_API명세서_v2.0.xlsx")
ws2 = wb["API 상세 명세"]

# ── 스타일 ──
DARK="2C2C2C"; WHITE="FFFFFF"; BLUE="2980B9"; GREEN="27AE60"; RED="E74C3C"; ORANGE="E67E22"
LIGHT_GRAY="F5F5F5"
hf=Font(name="맑은 고딕",bold=True,size=11,color=WHITE)
hfill=PatternFill(start_color=DARK,end_color=DARK,fill_type="solid")
bf=Font(name="맑은 고딕",size=10)
wa=Alignment(wrap_text=True,vertical="top",horizontal="left")
ca=Alignment(wrap_text=True,vertical="center",horizontal="center")
tb=Border(left=Side("thin","CCCCCC"),right=Side("thin","CCCCCC"),top=Side("thin","CCCCCC"),bottom=Side("thin","CCCCCC"))
af=PatternFill(start_color=LIGHT_GRAY,end_color=LIGHT_GRAY,fill_type="solid")
MF={
    "GET":Font(name="맑은 고딕",size=10,bold=True,color=BLUE),
    "POST":Font(name="맑은 고딕",size=10,bold=True,color=GREEN),
    "PUT":Font(name="맑은 고딕",size=10,bold=True,color=ORANGE),
    "PATCH":Font(name="맑은 고딕",size=10,bold=True,color=ORANGE),
    "DELETE":Font(name="맑은 고딕",size=10,bold=True,color=RED),
}

H2=["API ID","Method","엔드포인트","기능명","Path Params","Query Params",
    "Request Body","Request 예시","Response Body","Response 예시",
    "상태코드","에러 케이스","비고"]
W2=[14,8,38,15,22,24,42,42,42,48,20,32,24]

for c,(h,w) in enumerate(zip(H2,W2),1):
    cell=ws2.cell(row=1,column=c,value=h); cell.font=hf; cell.fill=hfill; cell.alignment=ca; cell.border=tb
    ws2.column_dimensions[get_column_letter(c)].width=w
ws2.freeze_panes="A2"

# ── 전체 90개 API 상세 데이터 ──
# 형식: [API_ID, Method, URI, 기능명, PathParam, QueryParam, ReqBody, ReqExample, ResBody, ResExample, StatusCodes, ErrorCases, Notes]

D = [
# ===== 인증 (AUTH) 10개 =====
["AUTH-API-001","POST","/api/v1/auth/login","로그인",
 "-","-",
 '{\n "email": "string (필수)",\n "password": "string (필수)"\n}',
 '{\n "email": "smith@example.com",\n "password": "mypass123"\n}',
 '{\n "accessToken": "string",\n "refreshToken": "string",\n "tokenType": "Bearer",\n "expiresIn": 3600,\n "user": {\n  "userId": "long",\n  "name": "string",\n  "email": "string",\n  "profileImageUrl": "string|null",\n  "role": "USER|ADMIN"\n }\n}',
 '{\n "accessToken": "eyJhbG...",\n "refreshToken": "dGhpcyBp...",\n "tokenType": "Bearer",\n "expiresIn": 3600,\n "user": {\n  "userId": 1,\n  "name": "김스미스",\n  "email": "smith@example.com",\n  "profileImageUrl": null,\n  "role": "USER"\n }\n}',
 "200 OK\n400 Bad Request\n401 Unauthorized\n403 Forbidden",
 "이메일 미존재 → 401\n비밀번호 불일치 → 401\n이메일 형식 오류 → 400\n비활성 계정 → 403",
 "중복 클릭 방지\nRefresh Token httpOnly cookie 권장"],

["AUTH-API-002","POST","/api/v1/auth/token/refresh","토큰 갱신",
 "-","-",
 '{"refreshToken": "string (필수)"}',
 '{"refreshToken": "dGhpcyBp..."}',
 '{"accessToken": "string", "refreshToken": "string", "expiresIn": 3600}',
 '{"accessToken": "eyJ..new", "refreshToken": "new...", "expiresIn": 3600}',
 "200 OK\n401 Unauthorized",
 "Refresh Token 만료 → 401\n무효 토큰 → 401",
 "Refresh Token Rotation 적용"],

["AUTH-API-003","POST","/api/v1/auth/token/verify","토큰 유효성 검증",
 "-","-",
 "-",
 "POST /api/v1/auth/token/verify\nAuthorization: Bearer eyJ...",
 '{"valid": "boolean", "userId": "long", "expiresAt": "datetime"}',
 '{"valid": true, "userId": 1, "expiresAt": "2026-02-17T15:00:00"}',
 "200 OK\n401 Unauthorized",
 "만료 토큰 → 401\n위변조 토큰 → 401",
 "앱 시작(SPL-003) 시 호출\nvalid=false면 refresh 시도"],

["AUTH-API-004","POST","/api/v1/auth/logout","로그아웃",
 "-","-",
 '{"refreshToken": "string (필수)"}',
 '{"refreshToken": "dGhpcyBp..."}',
 '{"message": "로그아웃 완료"}',
 '{"message": "로그아웃 완료"}',
 "200 OK\n401 Unauthorized",
 "이미 만료된 토큰 → 200 (멱등성)",
 "서버 Refresh Token 블랙리스트\nDEV-API-002 FCM 삭제도 함께 호출"],

["AUTH-API-005","POST","/api/v1/auth/signup","회원가입",
 "-","-",
 '{\n "name": "string (필수)",\n "phone": "string (필수, 010-XXXX-XXXX)",\n "email": "string (필수)",\n "password": "string (필수, 6자 이상)"\n}',
 '{\n "name": "김스미스",\n "phone": "010-1234-5678",\n "email": "smith@example.com",\n "password": "mypass123"\n}',
 '{\n "userId": "long",\n "name": "string",\n "email": "string",\n "createdAt": "datetime"\n}',
 '{\n "userId": 1,\n "name": "김스미스",\n "email": "smith@example.com",\n "createdAt": "2026-02-17T10:00:00"\n}',
 "201 Created\n400 Bad Request\n409 Conflict",
 "이메일 중복 → 409\n전화번호 중복 → 409\n비밀번호 6자 미만 → 400\n필수값 누락 → 400",
 "bcrypt 해싱 저장\nuser_settings 자동 생성"],

["AUTH-API-006","GET","/api/v1/auth/check-email","이메일 중복 확인",
 "-","email: string (필수)",
 "-",
 "GET /api/v1/auth/check-email?email=smith@example.com",
 '{"available": "boolean", "message": "string"}',
 '{"available": true, "message": "사용 가능한 이메일입니다."}',
 "200 OK\n400 Bad Request",
 "이메일 형식 오류 → 400",
 "실시간 검증 (디바운스 300ms 권장)"],

["AUTH-API-007","GET","/api/v1/auth/check-phone","전화번호 중복 확인",
 "-","phone: string (필수, 010-XXXX-XXXX)",
 "-",
 "GET /api/v1/auth/check-phone?phone=010-1234-5678",
 '{"available": "boolean", "message": "string"}',
 '{"available": false, "message": "이미 등록된 전화번호입니다."}',
 "200 OK\n400 Bad Request",
 "형식 오류 → 400",
 "실시간 검증 (디바운스 300ms 권장)"],

["AUTH-API-008","POST","/api/v1/auth/find-email","아이디 찾기",
 "-","-",
 '{"phone": "string (필수)"}',
 '{"phone": "010-1234-5678"}',
 '{"email": "string (마스킹)", "createdAt": "date"}',
 '{"email": "sm***@example.com", "createdAt": "2026-01-15"}',
 "200 OK\n404 Not Found",
 "미등록 전화번호 → 404",
 "이메일 부분 마스킹 처리"],

["AUTH-API-009","POST","/api/v1/auth/reset-password/request","비밀번호 재설정 요청",
 "-","-",
 '{"email": "string (필수)"}',
 '{"email": "smith@example.com"}',
 '{"message": "string", "expiresIn": 300}',
 '{"message": "재설정 코드가 이메일로 발송되었습니다.", "expiresIn": 300}',
 "200 OK\n404 Not Found\n429 Too Many",
 "미등록 이메일 → 404\n연속 요청 → 429",
 "5분 유효 코드\nRate Limit 분당 3회"],

["AUTH-API-010","POST","/api/v1/auth/reset-password/confirm","비밀번호 재설정 확인",
 "-","-",
 '{\n "email": "string (필수)",\n "code": "string (필수)",\n "newPassword": "string (필수, 6자 이상)"\n}',
 '{\n "email": "smith@example.com",\n "code": "A3F82K",\n "newPassword": "newpass456"\n}',
 '{"message": "비밀번호가 변경되었습니다."}',
 '{"message": "비밀번호가 변경되었습니다."}',
 "200 OK\n400 Bad Request\n401 Unauthorized",
 "잘못된 코드 → 401\n만료 코드 → 401\n비밀번호 규칙 미충족 → 400",
 "코드 사용 후 즉시 무효화"],

# ===== 홈 (HOME) 2개 =====
["HOME-API-001","GET","/api/v1/home/summary","홈 요약 조회",
 "-","-","-",
 "GET /api/v1/home/summary\nAuthorization: Bearer eyJ...",
 '{\n "nextReservation": {\n  "reservationId": "long|null",\n  "startTime": "HH:mm",\n  "endTime": "HH:mm",\n  "facilityName": "string"\n },\n "congestion": {\n  "status": "LOW|MEDIUM|HIGH",\n  "label": "string"\n },\n "membership": {\n  "type": "string",\n  "dDay": "int",\n  "status": "ACTIVE|EXPIRED|null"\n }\n}',
 '{\n "nextReservation": {\n  "reservationId": 42,\n  "startTime": "09:00",\n  "endTime": "10:30",\n  "facilityName": "메인 피트니스 존"\n },\n "congestion": {"status": "LOW", "label": "원활합니다"},\n "membership": {"type": "3개월 정기권", "dDay": 45, "status": "ACTIVE"}\n}',
 "200 OK\n401 Unauthorized",
 "예약 없음 → nextReservation: null\n회원권 없음 → membership: null",
 "Pull-to-Refresh 대응"],

["HOME-API-002","GET","/api/v1/home/announcements","홈 공지 미리보기",
 "-","limit: int (기본 5)",
 "-",
 "GET /api/v1/home/announcements?limit=5",
 '{"announcements": [{"announcementId": "long", "tag": "NOTICE|EVENT", "title": "string", "summary": "string", "publishedAt": "datetime"}]}',
 '{"announcements": [{"announcementId": 1, "tag": "NOTICE", "title": "설날 연휴 운영 시간 안내", "summary": "2월 8일~10일...", "publishedAt": "2026-02-07T09:00:00"}]}',
 "200 OK","빈 목록 → announcements: []","가로 스크롤 카드용"],

# ===== 사용자 (USER) 5개 =====
["USER-API-001","GET","/api/v1/users/me","내 정보 조회",
 "-","-","-",
 "GET /api/v1/users/me\nAuthorization: Bearer eyJ...",
 '{\n "userId": "long", "name": "string",\n "email": "string", "phone": "string",\n "profileImageUrl": "string|null",\n "membership": {"type": "string", "dDay": "int", "status": "string"},\n "attendanceRate": "double (%)",\n "todayVolume": "double (kg)"\n}',
 '{\n "userId": 1, "name": "김스미스",\n "email": "smith@example.com",\n "phone": "010-1234-5678",\n "profileImageUrl": null,\n "membership": {"type": "3개월 정기권", "dDay": 42, "status": "ACTIVE"},\n "attendanceRate": 92.0, "todayVolume": 1124.0\n}',
 "200 OK\n401 Unauthorized","미인증 → 401","내정보 화면 통합 데이터"],

["USER-API-002","PUT","/api/v1/users/me","내 정보 수정",
 "-","-",
 '{\n "name": "string (선택)",\n "phone": "string (선택)",\n "profileImageUrl": "string (선택)"\n}',
 '{"name": "김스미스", "phone": "010-9876-5432"}',
 '{\n "userId": "long", "name": "string",\n "phone": "string", "profileImageUrl": "string|null",\n "updatedAt": "datetime"\n}',
 '{"userId": 1, "name": "김스미스", "phone": "010-9876-5432", "profileImageUrl": null, "updatedAt": "2026-02-17T15:00:00"}',
 "200 OK\n400 Bad Request\n409 Conflict",
 "전화번호 중복 → 409\n형식 오류 → 400",
 "변경된 필드만 전달"],

["USER-API-003","POST","/api/v1/users/me/profile-image","프로필 이미지 업로드",
 "-","-",
 "multipart/form-data\nfile: 이미지 파일 (필수, JPG/PNG/WEBP, 최대 5MB)",
 "Content-Type: multipart/form-data\n[file binary data]",
 '{"profileImageUrl": "string"}',
 '{"profileImageUrl": "https://cdn.smithlife.co.kr/profiles/1/abc123.jpg"}',
 "200 OK\n413 Payload Too Large\n415 Unsupported Media Type",
 "5MB 초과 → 413\n미지원 형식 → 415",
 "기존 이미지 자동 덮어쓰기\nCDN URL 반환"],

["USER-API-004","PUT","/api/v1/users/me/password","비밀번호 변경",
 "-","-",
 '{\n "currentPassword": "string (필수)",\n "newPassword": "string (필수, 6자 이상)"\n}',
 '{"currentPassword": "mypass123", "newPassword": "newpass456"}',
 '{"message": "비밀번호가 변경되었습니다."}',
 '{"message": "비밀번호가 변경되었습니다."}',
 "200 OK\n400 Bad Request",
 "현재 비밀번호 불일치 → 400(INVALID_PASSWORD)\n새 비밀번호 6자 미만 → 400",
 "변경 후 기존 세션 유지"],

["USER-API-005","DELETE","/api/v1/users/me","회원 탈퇴",
 "-","-",
 '{"password": "string (필수)"}',
 '{"password": "mypass123"}',
 '{"message": "회원 탈퇴가 완료되었습니다."}',
 '{"message": "회원 탈퇴가 완료되었습니다."}',
 "200 OK\n400 Bad Request\n401 Unauthorized",
 "비밀번호 불일치 → 400",
 "Soft delete (is_active=false)\n토큰 무효화 → 로그인 화면 이동"],

# ===== 회원권 (MEMBERSHIP) 3개 =====
["MBSP-API-001","GET","/api/v1/memberships/me","내 회원권 조회",
 "-","-","-",
 "GET /api/v1/memberships/me",
 '{\n "membershipId": "long|null",\n "type": "string",\n "startDate": "date",\n "endDate": "date",\n "dDay": "int",\n "status": "ACTIVE|EXPIRED|PAUSED|CANCELLED"\n}',
 '{"membershipId": 1, "type": "3개월 정기권", "startDate": "2026-01-01", "endDate": "2026-03-31", "dDay": 42, "status": "ACTIVE"}',
 "200 OK","회원권 없음 → null 객체 반환","D-day 서버 계산"],

["MBSP-API-002","POST","/api/v1/memberships","회원권 등록/연장",
 "-","-",
 '{"type": "string (필수, 1개월|3개월|6개월|12개월)"}',
 '{"type": "3개월"}',
 '{\n "membershipId": "long", "type": "string",\n "startDate": "date", "endDate": "date",\n "status": "ACTIVE"\n}',
 '{"membershipId": 2, "type": "3개월", "startDate": "2026-04-01", "endDate": "2026-06-30", "status": "ACTIVE"}',
 "201 Created\n400 Bad Request",
 "잘못된 type → 400",
 "기존 만료 회원권 있으면 연장 처리"],

["MBSP-API-003","PATCH","/api/v1/memberships/{membershipId}/pause","회원권 일시정지",
 "membershipId: long","-",
 '{"pauseDays": "int (필수, 1~30)"}',
 '{"pauseDays": 14}',
 '{"membershipId": "long", "status": "PAUSED", "pauseEndDate": "date"}',
 '{"membershipId": 1, "status": "PAUSED", "pauseEndDate": "2026-03-03"}',
 "200 OK\n400 Bad Request",
 "30일 초과 → 400\n이미 PAUSED → 400\n만료 회원권 → 400",
 "정지 기간만큼 만료일 연장"],

# ===== 출석/QR (ATTENDANCE) 5개 =====
["ATT-API-001","POST","/api/v1/attendance/qr-token","QR 토큰 생성",
 "-","-","-",
 "POST /api/v1/attendance/qr-token\nAuthorization: Bearer eyJ...",
 '{\n "qrToken": "string",\n "qrImageBase64": "string",\n "expiresAt": "datetime",\n "ttlSeconds": 60\n}',
 '{\n "qrToken": "SL-QR-a3f82k9d...",\n "qrImageBase64": "data:image/png;base64,...",\n "expiresAt": "2026-02-17T14:01:00",\n "ttlSeconds": 60\n}',
 "200 OK\n401 Unauthorized\n403 Forbidden",
 "회원권 만료 → 403\n비활성 회원 → 403",
 "1회성 + 60초 유효\n30초~1분 자동 갱신 호출"],

["ATT-API-002","POST","/api/v1/attendance/check-in","QR 체크인",
 "-","-",
 '{"qrToken": "string (필수)"}',
 '{"qrToken": "SL-QR-a3f82k9d..."}',
 '{\n "attendanceId": "long",\n "userId": "long",\n "checkInTime": "datetime",\n "status": "CHECKED_IN"\n}',
 '{"attendanceId": 10, "userId": 1, "checkInTime": "2026-02-17T09:00:00", "status": "CHECKED_IN"}',
 "201 Created\n400 Bad Request\n401 Unauthorized",
 "만료 QR → 400\n이미 사용된 QR → 400\n유효하지 않은 QR → 400",
 "키오스크/직원 전용 (Admin 권한)\n이미 체크인 상태면 무시"],

["ATT-API-003","PATCH","/api/v1/attendance/{attendanceId}/check-out","체크아웃",
 "attendanceId: long","-","-",
 "PATCH /api/v1/attendance/10/check-out",
 '{"attendanceId": "long", "checkOutTime": "datetime", "status": "CHECKED_OUT"}',
 '{"attendanceId": 10, "checkOutTime": "2026-02-17T11:30:00", "status": "CHECKED_OUT"}',
 "200 OK\n404 Not Found",
 "존재하지 않는 출석 → 404\n이미 체크아웃 → 400",
 "퇴장 시 자동 호출"],

["ATT-API-004","GET","/api/v1/attendance/me","내 출석 기록",
 "-","month: string (필수, yyyy-MM)",
 "-",
 "GET /api/v1/attendance/me?month=2026-02",
 '{\n "month": "string",\n "totalDays": "int",\n "attendedDays": "int",\n "attendanceRate": "double",\n "records": [{"date": "date", "checkInTime": "time", "checkOutTime": "time|null"}]\n}',
 '{\n "month": "2026-02",\n "totalDays": 18, "attendedDays": 16,\n "attendanceRate": 88.9,\n "records": [{"date": "2026-02-17", "checkInTime": "09:00", "checkOutTime": "11:30"}]\n}',
 "200 OK","출석 없음 → records: []","참석율 자동 계산"],

["ATT-API-005","GET","/api/v1/attendance/me/status","현재 출석 상태",
 "-","-","-",
 "GET /api/v1/attendance/me/status",
 '{"isCheckedIn": "boolean", "attendanceId": "long|null", "checkInTime": "datetime|null"}',
 '{"isCheckedIn": true, "attendanceId": 10, "checkInTime": "2026-02-17T09:00:00"}',
 "200 OK","미체크인 → isCheckedIn: false","QR 모달 상태 표시용"],

# ===== 시설 (FACILITY) 2개 =====
["FAC-API-001","GET","/api/v1/facilities","시설 목록",
 "-","-","-",
 "GET /api/v1/facilities",
 '{"facilities": [{"facilityId": "long", "name": "string", "description": "string|null", "maxCapacity": "int", "isActive": "boolean"}]}',
 '{"facilities": [{"facilityId": 1, "name": "메인 피트니스 존", "description": "메인 웨이트/유산소", "maxCapacity": 50, "isActive": true}]}',
 "200 OK","-","캐싱 권장 (변경 빈도 낮음)"],

["FAC-API-002","GET","/api/v1/facilities/{facilityId}/congestion","시설 혼잡도",
 "facilityId: long","-","-",
 "GET /api/v1/facilities/1/congestion",
 '{"facilityId": "long", "facilityName": "string", "currentCount": "int", "maxCapacity": "int", "congestion": "LOW|MEDIUM|HIGH", "label": "string"}',
 '{"facilityId": 1, "facilityName": "메인 피트니스 존", "currentCount": 15, "maxCapacity": 50, "congestion": "LOW", "label": "원활합니다"}',
 "200 OK\n404 Not Found",
 "존재하지 않는 시설 → 404",
 "LOW: ~40%, MEDIUM: 40~80%, HIGH: 80%~"],

# ===== 시간 슬롯 1개 =====
["SLOT-API-001","GET","/api/v1/facilities/{facilityId}/slots","시간 슬롯 조회",
 "facilityId: long","date: string (필수, yyyy-MM-dd)",
 "-",
 "GET /api/v1/facilities/1/slots?date=2026-02-10",
 '{\n "facilityName": "string", "date": "date",\n "slots": [{\n  "slotId": "long",\n  "startTime": "HH:mm", "endTime": "HH:mm",\n  "maxCapacity": "int", "currentCount": "int",\n  "congestion": "LOW|MEDIUM|HIGH",\n  "myReserved": "boolean"\n }]\n}',
 '{\n "facilityName": "메인 피트니스 존",\n "date": "2026-02-10",\n "slots": [\n  {"slotId": 101, "startTime": "09:00", "endTime": "10:00", "maxCapacity": 50, "currentCount": 12, "congestion": "LOW", "myReserved": false},\n  {"slotId": 102, "startTime": "10:00", "endTime": "11:00", "maxCapacity": 50, "currentCount": 35, "congestion": "MEDIUM", "myReserved": true}\n ]\n}',
 "200 OK\n404 Not Found",
 "미존재 시설 → 404",
 "myReserved로 취소 가능 판단"],

# ===== 예약 (RESERVATION) 6개 =====
["RESV-API-001","POST","/api/v1/reservations","예약 생성",
 "-","-",
 '{"slotIds": ["long[] (필수, 1개 이상)"]}',
 '{"slotIds": [101, 102]}',
 '{\n "reservations": [{\n  "reservationId": "long",\n  "reservationNo": "string",\n  "slotId": "long",\n  "date": "date",\n  "startTime": "HH:mm",\n  "endTime": "HH:mm",\n  "facilityName": "string",\n  "status": "CONFIRMED"\n }]\n}',
 '{"reservations": [{"reservationId": 42, "reservationNo": "#8812", "slotId": 101, "date": "2026-02-10", "startTime": "09:00", "endTime": "10:00", "facilityName": "메인 피트니스 존", "status": "CONFIRMED"}]}',
 "201 Created\n400 Bad Request\n409 Conflict",
 "만석 → 409(SLOT_FULL)\n중복 예약 → 409\n빈 slotIds → 400",
 "복수 시간 동시 예약\ncurrent_count 증가\n알림 자동 생성"],

["RESV-API-002","GET","/api/v1/reservations/me","내 예약 목록",
 "-","date: string (선택, yyyy-MM-dd)\nfacilityId: long (선택)",
 "-",
 "GET /api/v1/reservations/me?date=2026-02-10",
 '{"reservations": [{"reservationId": "long", "reservationNo": "string", "date": "date", "startTime": "HH:mm", "endTime": "HH:mm", "facilityName": "string", "status": "CONFIRMED|CANCELLED|COMPLETED"}]}',
 '{"reservations": [{"reservationId": 42, "reservationNo": "#8812", "date": "2026-02-10", "startTime": "09:00", "endTime": "10:00", "facilityName": "메인 피트니스 존", "status": "CONFIRMED"}]}',
 "200 OK","예약 없음 → reservations: []","취소 모달용 (CONFIRMED만 필터)"],

["RESV-API-003","GET","/api/v1/reservations/me/upcoming","다음 예약",
 "-","-","-",
 "GET /api/v1/reservations/me/upcoming",
 '{"reservationId": "long|null", "reservationNo": "string", "date": "date", "startTime": "HH:mm", "endTime": "HH:mm", "facilityName": "string"}',
 '{"reservationId": 42, "reservationNo": "#8812", "date": "2026-02-10", "startTime": "09:00", "endTime": "10:30", "facilityName": "메인 피트니스 존"}',
 "200 OK","예약 없음 → null","홈 다음 예약 카드용"],

["RESV-API-004","PATCH","/api/v1/reservations/{reservationId}/cancel","예약 취소",
 "reservationId: long","-","-",
 "PATCH /api/v1/reservations/42/cancel",
 '{"reservationId": "long", "reservationNo": "string", "status": "CANCELLED", "cancelledAt": "datetime"}',
 '{"reservationId": 42, "reservationNo": "#8812", "status": "CANCELLED", "cancelledAt": "2026-02-10T08:30:00"}',
 "200 OK\n400 Bad Request\n403 Forbidden\n404 Not Found",
 "이미 취소 → 400\n타인 예약 → 403\n완료 예약 → 400",
 "취소 전 확인 다이얼로그\ncurrent_count 감소"],

["RESV-API-005","GET","/api/v1/reservations/{reservationId}","예약 상세",
 "reservationId: long","-","-",
 "GET /api/v1/reservations/42",
 '{\n "reservationId": "long", "reservationNo": "string",\n "date": "date", "startTime": "HH:mm", "endTime": "HH:mm",\n "facilityName": "string", "status": "string",\n "reservedAt": "datetime", "cancelledAt": "datetime|null"\n}',
 '{"reservationId": 42, "reservationNo": "#8812", "date": "2026-02-10", "startTime": "09:00", "endTime": "10:00", "facilityName": "메인 피트니스 존", "status": "CONFIRMED", "reservedAt": "2026-02-09T20:00:00", "cancelledAt": null}',
 "200 OK\n404 Not Found","미존재 → 404","알림 클릭 → 예약 상세 이동"],

["RESV-API-006","GET","/api/v1/reservations/me/calendar","월별 예약 캘린더",
 "-","month: string (필수, yyyy-MM)",
 "-",
 "GET /api/v1/reservations/me/calendar?month=2026-02",
 '{"month": "string", "dates": [{"date": "date", "count": "int"}]}',
 '{"month": "2026-02", "dates": [{"date": "2026-02-10", "count": 2}, {"date": "2026-02-12", "count": 1}]}',
 "200 OK","예약 없는 달 → dates: []","캘린더 dot 마커 표시용"],

# ===== 이용내역 (USAGE) 2개 =====
["USAGE-API-001","GET","/api/v1/usage-history/me","이용내역 목록",
 "-","page: int (기본 0)\nsize: int (기본 20)\nstatus: string (선택, COMPLETED|CANCELLED|NO_SHOW)\nfrom: date (선택)\nto: date (선택)",
 "-",
 "GET /api/v1/usage-history/me?page=0&size=20",
 '{\n "content": [{\n  "usageId": "long", "facilityName": "string",\n  "usageDate": "date",\n  "startTime": "HH:mm", "endTime": "HH:mm",\n  "reservationNo": "string|null",\n  "status": "COMPLETED|CANCELLED|NO_SHOW"\n }],\n "totalElements": "long",\n "totalPages": "int", "hasNext": "boolean"\n}',
 '{"content": [{"usageId": 1, "facilityName": "시설 이용", "usageDate": "2026-01-28", "startTime": "09:00", "endTime": "11:00", "reservationNo": "#8812", "status": "COMPLETED"}], "totalElements": 15, "totalPages": 1, "hasNext": false}',
 "200 OK","빈 목록 → content: []","최신순 정렬\n무한 스크롤"],

["USAGE-API-002","GET","/api/v1/usage-history/{usageId}","이용내역 상세",
 "usageId: long","-","-",
 "GET /api/v1/usage-history/1",
 '{\n "usageId": "long", "facilityName": "string",\n "usageDate": "date", "startTime": "HH:mm", "endTime": "HH:mm",\n "reservationId": "long|null", "reservationNo": "string|null",\n "status": "string", "createdAt": "datetime"\n}',
 '{"usageId": 1, "facilityName": "시설 이용", "usageDate": "2026-01-28", "startTime": "09:00", "endTime": "11:00", "reservationId": 42, "reservationNo": "#8812", "status": "COMPLETED", "createdAt": "2026-01-28T11:00:00"}',
 "200 OK\n404 Not Found","미존재 → 404","-"],

# ===== 운동 종목 (EXERCISE) 3개 =====
["EXER-API-001","GET","/api/v1/exercises","운동 종목 검색",
 "-","bodyPart: string (선택)\nequipment: string (선택)\nkeyword: string (선택)",
 "-",
 "GET /api/v1/exercises?bodyPart=가슴&keyword=벤치",
 '{"exercises": [{"exerciseId": "long", "name": "string", "bodyPart": "string", "equipment": "string|null", "imageUrl": "string|null"}]}',
 '{"exercises": [{"exerciseId": 1, "name": "벤치 프레스", "bodyPart": "가슴", "equipment": "바벨", "imageUrl": null}]}',
 "200 OK","결과 없음 → exercises: []","운동 추가/루틴 구성 시 사용"],

["EXER-API-002","GET","/api/v1/exercises/{exerciseId}","운동 종목 상세",
 "exerciseId: long","-","-",
 "GET /api/v1/exercises/1",
 '{\n "exerciseId": "long", "name": "string",\n "bodyPart": "string", "equipment": "string|null",\n "imageUrl": "string|null", "description": "string|null"\n}',
 '{"exerciseId": 1, "name": "벤치 프레스", "bodyPart": "가슴", "equipment": "바벨", "imageUrl": null, "description": "가슴 근육 발달을 위한 기본 운동"}',
 "200 OK\n404 Not Found","미존재 → 404","운동 상세 화면 헤더 데이터"],

["EXER-API-003","GET","/api/v1/exercises/{exerciseId}/history","운동 이전 기록",
 "exerciseId: long","limit: int (기본 10)",
 "-",
 "GET /api/v1/exercises/1/history?limit=5",
 '{"exerciseId": "long", "exerciseName": "string", "history": [{"date": "date", "bestSetWeightKg": "double", "bestSetReps": "int", "totalVolume": "double"}]}',
 '{"exerciseId": 1, "exerciseName": "벤치 프레스", "history": [{"date": "2026-02-15", "bestSetWeightKg": 70.0, "bestSetReps": 10, "totalVolume": 2100.0}]}',
 "200 OK\n404 Not Found","기록 없음 → history: []","운동 상세 지난 기록 참조용"],

# ===== 루틴 (ROUTINE) 7개 =====
["RTN-API-001","GET","/api/v1/routines/recommended","추천 루틴 목록",
 "-","-","-",
 "GET /api/v1/routines/recommended",
 '{"routines": [{"routineId": "long", "name": "string", "goal": "string", "difficulty": "string", "estimatedMin": "int", "frequency": "string", "imageUrl": "string|null"}]}',
 '{"routines": [{"routineId": 1, "name": "파워 빌딩", "goal": "MUSCLE_GAIN", "difficulty": "BEGINNER", "estimatedMin": 45, "frequency": "주 3회", "imageUrl": null}]}',
 "200 OK","-","8개 추천 루틴\n가로 스크롤/그리드용"],

["RTN-API-002","GET","/api/v1/routines","루틴 검색",
 "-","keyword: string (선택)\ngoal: MUSCLE_GAIN|DIET|STAMINA (선택)\ndifficulty: BEGINNER|INTERMEDIATE|ADVANCED (선택)",
 "-",
 "GET /api/v1/routines?keyword=파워&goal=MUSCLE_GAIN",
 '{"routines": [{"routineId": "long", "name": "string", "goal": "string", "difficulty": "string", "estimatedMin": "int", "frequency": "string", "isPublic": "boolean", "creatorName": "string|null"}]}',
 '{"routines": [{"routineId": 1, "name": "파워 빌딩", "goal": "MUSCLE_GAIN", "difficulty": "BEGINNER", "estimatedMin": 45, "frequency": "주 3회", "isPublic": true, "creatorName": null}]}',
 "200 OK","결과 없음 → routines: []","공개 루틴 + 추천 루틴 통합 검색"],

["RTN-API-003","GET","/api/v1/routines/{routineId}","루틴 상세",
 "routineId: long","-","-",
 "GET /api/v1/routines/1",
 '{\n "routineId": "long", "name": "string",\n "goal": "MUSCLE_GAIN|DIET|STAMINA",\n "difficulty": "BEGINNER|INTERMEDIATE|ADVANCED",\n "estimatedMin": "int", "frequency": "string",\n "description": "string", "imageUrl": "string|null",\n "isRecommended": "boolean", "isFavorited": "boolean",\n "exerciseCount": "int",\n "exercises": [{\n  "exerciseId": "long", "name": "string",\n  "bodyPart": "string", "equipment": "string",\n  "imageUrl": "string|null",\n  "targetSets": "int",\n  "targetRepsMin": "int", "targetRepsMax": "int|null",\n  "orderIndex": "int"\n }]\n}',
 '{\n "routineId": 1, "name": "초보자 파워 빌딩",\n "goal": "MUSCLE_GAIN", "difficulty": "BEGINNER",\n "estimatedMin": 45, "frequency": "주 3회",\n "description": "기초 근력 향상...",\n "imageUrl": "/images/routine1.jpg",\n "isRecommended": true, "isFavorited": false,\n "exerciseCount": 5,\n "exercises": [{"exerciseId": 1, "name": "벤치 프레스", "bodyPart": "가슴", "equipment": "바벨", "imageUrl": null, "targetSets": 3, "targetRepsMin": 10, "targetRepsMax": 12, "orderIndex": 1}]\n}',
 "200 OK\n404 Not Found","미존재 → 404","isFavorited: 로그인 사용자 기준"],

["RTN-API-004","POST","/api/v1/routines","루틴 생성",
 "-","-",
 '{\n "name": "string (필수)",\n "goal": "MUSCLE_GAIN|DIET|STAMINA (필수)",\n "estimatedMin": "int (필수)",\n "frequency": "string (선택)",\n "isPublic": "boolean (기본 false)",\n "exercises": [{\n  "exerciseId": "long (필수)",\n  "orderIndex": "int (필수)",\n  "targetSets": "int (필수)",\n  "targetRepsMin": "int (필수)",\n  "targetRepsMax": "int (선택)"\n }]\n}',
 '{\n "name": "월요일 하체 루틴",\n "goal": "MUSCLE_GAIN",\n "estimatedMin": 40,\n "frequency": "주 2회",\n "isPublic": false,\n "exercises": [{"exerciseId": 2, "orderIndex": 1, "targetSets": 4, "targetRepsMin": 8, "targetRepsMax": 10}]\n}',
 '{"routineId": "long", "name": "string", "createdAt": "datetime"}',
 '{"routineId": 9, "name": "월요일 하체 루틴", "createdAt": "2026-02-17T15:00:00"}',
 "201 Created\n400 Bad Request",
 "이름 누락 → 400\n운동 0개 → 400\n목표 미선택 → 400",
 "exercises 배열로 순서+세트 한번에 저장"],

["RTN-API-005","PUT","/api/v1/routines/{routineId}","루틴 수정",
 "routineId: long","-",
 '{\n "name": "string (선택)",\n "goal": "string (선택)",\n "isPublic": "boolean (선택)",\n "exercises": [동일 구조] (선택)\n}',
 '{"name": "월요일 하체 루틴 v2", "exercises": [{"exerciseId": 2, "orderIndex": 1, "targetSets": 5, "targetRepsMin": 6, "targetRepsMax": 8}]}',
 '{"routineId": "long", "name": "string", "updatedAt": "datetime"}',
 '{"routineId": 9, "name": "월요일 하체 루틴 v2", "updatedAt": "2026-02-17T16:00:00"}',
 "200 OK\n403 Forbidden\n404 Not Found",
 "타인 루틴 → 403\n추천 루틴 수정 → 403",
 "exercises 전달 시 기존 구성 대체(PUT)\n순서 변경 = exercises 재전달"],

["RTN-API-006","DELETE","/api/v1/routines/{routineId}","루틴 삭제",
 "routineId: long","-","-",
 "DELETE /api/v1/routines/9",
 '{"message": "루틴이 삭제되었습니다."}',
 '{"message": "루틴이 삭제되었습니다."}',
 "200 OK\n403 Forbidden\n404 Not Found",
 "타인 루틴 → 403\n추천 루틴 삭제 → 403",
 "연관 favorite_routine도 CASCADE 삭제"],

["RTN-API-007","GET","/api/v1/routines/me","내 루틴 목록",
 "-","-","-",
 "GET /api/v1/routines/me",
 '{"routines": [{"routineId": "long", "name": "string", "goal": "string", "estimatedMin": "int", "exerciseCount": "int", "isPublic": "boolean", "createdAt": "datetime"}]}',
 '{"routines": [{"routineId": 9, "name": "월요일 하체 루틴", "goal": "MUSCLE_GAIN", "estimatedMin": 40, "exerciseCount": 3, "isPublic": false, "createdAt": "2026-02-17T15:00:00"}]}',
 "200 OK","루틴 없음 → routines: []","내가 생성한 루틴만"],

# ===== 찜 (FAVORITE) 3개 =====
["FAV-API-001","POST","/api/v1/routines/{routineId}/favorite","찜 추가",
 "routineId: long","-","-",
 "POST /api/v1/routines/1/favorite",
 '{"message": "찜 목록에 추가되었습니다."}',
 '{"message": "찜 목록에 추가되었습니다."}',
 "201 Created\n409 Conflict","이미 찜함 → 409","하트 채우기 애니메이션"],

["FAV-API-002","DELETE","/api/v1/routines/{routineId}/favorite","찜 해제",
 "routineId: long","-","-",
 "DELETE /api/v1/routines/1/favorite",
 '{"message": "찜 목록에서 제거되었습니다."}',
 '{"message": "찜 목록에서 제거되었습니다."}',
 "200 OK\n404 Not Found","찜하지 않은 루틴 → 404","하트 비우기"],

["FAV-API-003","GET","/api/v1/routines/favorites","찜한 루틴 목록",
 "-","-","-",
 "GET /api/v1/routines/favorites",
 '{"routines": [{"routineId": "long", "name": "string", "goal": "string", "difficulty": "string", "estimatedMin": "int", "frequency": "string"}]}',
 '{"routines": [{"routineId": 1, "name": "파워 빌딩", "goal": "MUSCLE_GAIN", "difficulty": "BEGINNER", "estimatedMin": 45, "frequency": "주 3회"}]}',
 "200 OK","찜 없음 → routines: []","-"],

# ===== 세션 (SESSION) 5개 =====
["SESS-API-001","POST","/api/v1/workout-sessions","세션 시작",
 "-","-",
 '{\n "routineId": "long|null",\n "sessionName": "string (필수)"\n}',
 '{"routineId": 1, "sessionName": "가슴 & 삼두 데이"}',
 '{\n "sessionId": "long", "sessionName": "string",\n "status": "ACTIVE", "startTime": "datetime",\n "exercises": [{"sessionExerciseId": "long", "exerciseId": "long", "exerciseName": "string", "bodyPart": "string", "orderIndex": "int", "targetSets": "int", "targetReps": "int", "status": "PENDING"}]\n}',
 '{"sessionId": 10, "sessionName": "가슴 & 삼두 데이", "status": "ACTIVE", "startTime": "2026-02-17T14:00:00", "exercises": [{"sessionExerciseId": 1, "exerciseId": 1, "exerciseName": "벤치 프레스", "bodyPart": "가슴", "orderIndex": 1, "targetSets": 3, "targetReps": 12, "status": "PENDING"}]}',
 "201 Created\n400 Bad Request\n409 Conflict",
 "활성 세션 존재 → 409\n세션명 미입력 → 400",
 "routineId 전달 시 운동 자동 복사\nACTIVE 세션 1개만 허용"],

["SESS-API-002","GET","/api/v1/workout-sessions/active","활성 세션 조회",
 "-","-","-",
 "GET /api/v1/workout-sessions/active",
 '{\n "sessionId": "long", "sessionName": "string",\n "status": "ACTIVE|PAUSED",\n "startTime": "datetime",\n "totalDurationSec": "int", "totalCalories": "double",\n "exercises": [{"sessionExerciseId": "long", "exerciseName": "string", "bodyPart": "string", "status": "PENDING|IN_PROGRESS|COMPLETED", "completedSets": "int", "targetSets": "int"}]\n}',
 '{"sessionId": 10, "sessionName": "가슴 & 삼두 데이", "status": "ACTIVE", "startTime": "2026-02-17T14:00:00", "totalDurationSec": 1455, "totalCalories": 320.0, "exercises": [{"sessionExerciseId": 1, "exerciseName": "벤치 프레스", "bodyPart": "가슴", "status": "COMPLETED", "completedSets": 4, "targetSets": 4}]}',
 "200 OK\n404 Not Found","활성 세션 없음 → 404","ACTIVE SESSION 카드 + 진행중 운동 리스트"],

["SESS-API-003","PATCH","/api/v1/workout-sessions/{sessionId}/end","세션 종료",
 "sessionId: long","-","-",
 "PATCH /api/v1/workout-sessions/10/end",
 '{\n "sessionId": "long", "status": "COMPLETED",\n "endTime": "datetime",\n "totalDurationSec": "int",\n "totalVolumeKg": "double",\n "totalCalories": "double",\n "reportId": "long"\n}',
 '{"sessionId": 10, "status": "COMPLETED", "endTime": "2026-02-17T14:45:12", "totalDurationSec": 2712, "totalVolumeKg": 4250.0, "totalCalories": 380.0, "reportId": 5}',
 "200 OK\n400 Bad Request\n404 Not Found",
 "이미 종료 → 400\n미존재 세션 → 404",
 "workout_report 자동 생성\npersonal_record 자동 갱신"],

["SESS-API-004","PATCH","/api/v1/workout-sessions/{sessionId}/pause","세션 일시정지",
 "sessionId: long","-","-",
 "PATCH /api/v1/workout-sessions/10/pause",
 '{"sessionId": "long", "status": "PAUSED"}',
 '{"sessionId": 10, "status": "PAUSED"}',
 "200 OK\n400 Bad Request","이미 PAUSED → 400\n종료된 세션 → 400","-"],

["SESS-API-005","PATCH","/api/v1/workout-sessions/{sessionId}/resume","세션 재개",
 "sessionId: long","-","-",
 "PATCH /api/v1/workout-sessions/10/resume",
 '{"sessionId": "long", "status": "ACTIVE"}',
 '{"sessionId": 10, "status": "ACTIVE"}',
 "200 OK\n400 Bad Request","이미 ACTIVE → 400\n종료된 세션 → 400","-"],

# ===== 세션 운동 (SE) 5개 =====
["SE-API-001","POST","/api/v1/workout-sessions/{sessionId}/exercises","세션 운동 추가",
 "sessionId: long","-",
 '{"exerciseId": "long (필수)", "targetSets": "int (기본 4)", "targetReps": "int (기본 12)"}',
 '{"exerciseId": 5, "targetSets": 3, "targetReps": 15}',
 '{"sessionExerciseId": "long", "exerciseId": "long", "exerciseName": "string", "bodyPart": "string", "orderIndex": "int", "status": "PENDING"}',
 '{"sessionExerciseId": 3, "exerciseId": 5, "exerciseName": "케이블 푸쉬다운", "bodyPart": "삼두", "orderIndex": 3, "status": "PENDING"}',
 "201 Created\n400 Bad Request\n404 Not Found",
 "종료된 세션 → 400\n미존재 운동 → 404",
 "orderIndex 자동 할당 (마지막+1)"],

["SE-API-002","GET","/api/v1/workout-sessions/{sessionId}/exercises","세션 운동 목록",
 "sessionId: long","-","-",
 "GET /api/v1/workout-sessions/10/exercises",
 '{"exercises": [{"sessionExerciseId": "long", "exerciseId": "long", "exerciseName": "string", "bodyPart": "string", "equipment": "string|null", "orderIndex": "int", "targetSets": "int", "targetReps": "int", "completedSets": "int", "status": "PENDING|IN_PROGRESS|COMPLETED"}]}',
 '{"exercises": [{"sessionExerciseId": 1, "exerciseId": 1, "exerciseName": "벤치 프레스", "bodyPart": "가슴", "equipment": "바벨", "orderIndex": 1, "targetSets": 4, "targetReps": 12, "completedSets": 4, "status": "COMPLETED"}]}',
 "200 OK\n404 Not Found","미존재 세션 → 404","진행중 운동 리스트 전체"],

["SE-API-003","PATCH","/api/v1/session-exercises/{seId}/complete","운동 완료",
 "seId: long","-","-",
 "PATCH /api/v1/session-exercises/1/complete",
 '{"sessionExerciseId": "long", "status": "COMPLETED"}',
 '{"sessionExerciseId": 1, "status": "COMPLETED"}',
 "200 OK\n400 Bad Request","이미 완료 → 400","운동 종료 버튼 클릭 시"],

["SE-API-004","PATCH","/api/v1/session-exercises/{seId}/start","운동 시작",
 "seId: long","-","-",
 "PATCH /api/v1/session-exercises/2/start",
 '{"sessionExerciseId": "long", "status": "IN_PROGRESS"}',
 '{"sessionExerciseId": 2, "status": "IN_PROGRESS"}',
 "200 OK\n400 Bad Request","이미 진행/완료 → 400","UP NEXT → 진행중"],

["SE-API-005","DELETE","/api/v1/session-exercises/{seId}","운동 삭제",
 "seId: long","-","-",
 "DELETE /api/v1/session-exercises/3",
 '{"message": "운동이 제거되었습니다."}',
 '{"message": "운동이 제거되었습니다."}',
 "200 OK\n404 Not Found","미존재 → 404","orderIndex 자동 재정렬"],

# ===== 세트 (SET) 5개 =====
["SET-API-001","POST","/api/v1/session-exercises/{seId}/sets","세트 추가",
 "seId: long","-",
 '{"weightKg": "double|null", "reps": "int|null"}',
 '{"weightKg": null, "reps": null}',
 '{"setId": "long", "setNumber": "int", "weightKg": "double|null", "reps": "int|null", "isCompleted": false}',
 '{"setId": 5, "setNumber": 5, "weightKg": null, "reps": null, "isCompleted": false}',
 "201 Created\n404 Not Found","미존재 운동 → 404","setNumber 자동 증가\n미입력 = '--' 표시"],

["SET-API-002","GET","/api/v1/session-exercises/{seId}/sets","세트 목록",
 "seId: long","-","-",
 "GET /api/v1/session-exercises/1/sets",
 '{"sets": [{"setId": "long", "setNumber": "int", "weightKg": "double|null", "reps": "int|null", "isCompleted": "boolean", "restTimeSec": "int|null", "completedAt": "datetime|null"}]}',
 '{"sets": [{"setId": 1, "setNumber": 1, "weightKg": 60.0, "reps": 12, "isCompleted": true, "restTimeSec": 84, "completedAt": "2026-02-17T14:05:00"}, {"setId": 2, "setNumber": 2, "weightKg": 60.0, "reps": 0, "isCompleted": false, "restTimeSec": null, "completedAt": null}]}',
 "200 OK\n404 Not Found","미존재 운동 → 404","세트 테이블 전체 데이터"],

["SET-API-003","PATCH","/api/v1/exercise-sets/{setId}","세트 수정",
 "setId: long","-",
 '{"weightKg": "double (선택)", "reps": "int (선택)", "restTimeSec": "int (선택)"}',
 '{"weightKg": 65.0, "reps": 10}',
 '{"setId": "long", "weightKg": "double", "reps": "int", "isCompleted": "boolean"}',
 '{"setId": 2, "weightKg": 65.0, "reps": 10, "isCompleted": false}',
 "200 OK\n404 Not Found","미존재 → 404","변경된 필드만 전달"],

["SET-API-004","PATCH","/api/v1/exercise-sets/{setId}/complete","세트 완료",
 "setId: long","-","-",
 "PATCH /api/v1/exercise-sets/2/complete",
 '{"setId": "long", "isCompleted": true, "completedAt": "datetime", "restTimerStarted": "boolean"}',
 '{"setId": 2, "isCompleted": true, "completedAt": "2026-02-17T14:08:24", "restTimerStarted": true}',
 "200 OK\n400 Bad Request","이미 완료 → 400\nKG/횟수 미입력 → 400","완료 시 PR 자동 비교/갱신\n휴식 타이머 자동 시작"],

["SET-API-005","DELETE","/api/v1/exercise-sets/{setId}","세트 삭제",
 "setId: long","-","-",
 "DELETE /api/v1/exercise-sets/5",
 '{"message": "세트가 삭제되었습니다."}',
 '{"message": "세트가 삭제되었습니다."}',
 "200 OK\n404 Not Found","미존재 → 404","setNumber 자동 재정렬"],

# ===== 개인기록 (PR) 2개 =====
["PR-API-001","GET","/api/v1/personal-records","운동별 최고 기록",
 "-","exerciseId: long (필수)",
 "-",
 "GET /api/v1/personal-records?exerciseId=1",
 '{"records": [{"recordId": "long", "recordType": "ONE_RM|MAX_VOLUME|MAX_REPS", "value": "double", "achievedAt": "date"}]}',
 '{"records": [{"recordId": 1, "recordType": "ONE_RM", "value": 75.0, "achievedAt": "2026-02-10"}]}',
 "200 OK","기록 없음 → records: []","지난 최고 기록(1RM) 카드용"],

["PR-API-002","GET","/api/v1/personal-records/me","전체 개인 기록",
 "-","-","-",
 "GET /api/v1/personal-records/me",
 '{"records": [{"exerciseId": "long", "exerciseName": "string", "bodyPart": "string", "recordType": "string", "value": "double", "achievedAt": "date"}]}',
 '{"records": [{"exerciseId": 1, "exerciseName": "벤치 프레스", "bodyPart": "가슴", "recordType": "ONE_RM", "value": 75.0, "achievedAt": "2026-02-10"}]}',
 "200 OK","기록 없음 → records: []","-"],

# ===== 리포트 (RPT) 4개 =====
["RPT-API-001","GET","/api/v1/workout-reports/{sessionId}","리포트 조회",
 "sessionId: long","-","-",
 "GET /api/v1/workout-reports/10",
 '{\n "reportId": "long", "reportDate": "date",\n "userName": "string",\n "totalTimeSec": "int", "totalTimeFormatted": "string",\n "totalVolumeKg": "double", "totalCalories": "double",\n "weeklyChangePct": "double|null",\n "motivationMsg": "string",\n "exercises": [{"exerciseName": "string", "sets": "int", "reps": "int", "weightKg": "double", "isCompleted": "boolean"}],\n "sharedAt": "datetime|null"\n}',
 '{\n "reportId": 5, "reportDate": "2026-02-17",\n "userName": "김스미스",\n "totalTimeSec": 2712, "totalTimeFormatted": "45:12",\n "totalVolumeKg": 4250.0, "totalCalories": 380.0,\n "weeklyChangePct": 12.0,\n "motivationMsg": "오늘도 한계를 넘으셨군요!",\n "exercises": [{"exerciseName": "백 스쿼트", "sets": 4, "reps": 10, "weightKg": 100.0, "isCompleted": true}],\n "sharedAt": null\n}',
 "200 OK\n404 Not Found","미완료 세션 → 404\n타인 세션 → 403","메달+동기부여 메시지"],

["RPT-API-002","GET","/api/v1/workout-reports/weekly","주간 차트",
 "-","date: string (선택, yyyy-MM-dd, 기본 오늘)",
 "-",
 "GET /api/v1/workout-reports/weekly?date=2026-02-17",
 '{\n "weekStart": "date", "weekEnd": "date",\n "weeklyChangePct": "double|null",\n "dailyData": [{"dayOfWeek": "string", "date": "date", "volumeKg": "double", "durationMin": "int"}]\n}',
 '{"weekStart": "2026-02-10", "weekEnd": "2026-02-16", "weeklyChangePct": 12.0, "dailyData": [{"dayOfWeek": "MON", "date": "2026-02-10", "volumeKg": 3200.0, "durationMin": 45}]}',
 "200 OK","운동 없는 주 → dailyData 0값","주간 운동량 바 차트용"],

["RPT-API-003","POST","/api/v1/workout-reports/{reportId}/share","리포트 공유",
 "reportId: long","-","-",
 "POST /api/v1/workout-reports/5/share",
 '{"shareImageUrl": "string", "sharedAt": "datetime"}',
 '{"shareImageUrl": "https://cdn.smithlife.co.kr/shares/rpt5.png", "sharedAt": "2026-02-17T15:30:00"}',
 "200 OK\n404 Not Found","미존재 리포트 → 404","이미지 생성 → CDN 업로드\nSNS 공유용 URL"],

["RPT-API-004","GET","/api/v1/workout-reports/today","오늘의 운동량",
 "-","-","-",
 "GET /api/v1/workout-reports/today",
 '{"date": "date", "totalVolumeKg": "double", "totalDurationSec": "int", "totalCalories": "double", "sessionCount": "int"}',
 '{"date": "2026-02-17", "totalVolumeKg": 1124.0, "totalDurationSec": 2712, "totalCalories": 380.0, "sessionCount": 1}',
 "200 OK","운동 없음 → 모두 0","내정보 오늘의 운동량 카드용"],

# ===== 설정 (SETT) 2개 =====
["SETT-API-001","GET","/api/v1/settings","설정 조회",
 "-","-","-",
 "GET /api/v1/settings",
 '{\n "notificationEnabled": "boolean",\n "darkMode": "SYSTEM|ON|OFF",\n "language": "ko|en|ja"\n}',
 '{"notificationEnabled": true, "darkMode": "SYSTEM", "language": "ko"}',
 "200 OK\n401 Unauthorized","-","설정 화면 초기 데이터"],

["SETT-API-002","PATCH","/api/v1/settings","설정 변경",
 "-","-",
 '{\n "notificationEnabled": "boolean (선택)",\n "darkMode": "SYSTEM|ON|OFF (선택)",\n "language": "ko|en|ja (선택)"\n}',
 '{"darkMode": "ON"}',
 '{"notificationEnabled": "boolean", "darkMode": "string", "language": "string", "updatedAt": "datetime"}',
 '{"notificationEnabled": true, "darkMode": "ON", "language": "ko", "updatedAt": "2026-02-17T15:00:00"}',
 "200 OK\n400 Bad Request",
 "잘못된 darkMode 값 → 400\n잘못된 language → 400",
 "PATCH: 전달된 필드만 변경\n즉시 반영 (SET-012)"],

# ===== 알림 (NOTI) 4개 =====
["NOTI-API-001","GET","/api/v1/notifications","알림 목록",
 "-","page: int (기본 0)\nsize: int (기본 20)\ntype: RESERVATION|MEMBERSHIP|WORKOUT|SYSTEM (선택)",
 "-",
 "GET /api/v1/notifications?page=0&size=20",
 '{\n "content": [{\n  "notificationId": "long",\n  "type": "RESERVATION|MEMBERSHIP|WORKOUT|SYSTEM",\n  "title": "string", "message": "string",\n  "subMessage": "string|null",\n  "isRead": "boolean",\n  "relatedUrl": "string|null",\n  "createdAt": "datetime",\n  "timeAgo": "string"\n }],\n "totalElements": "long", "totalPages": "int", "hasNext": "boolean"\n}',
 '{"content": [{"notificationId": 1, "type": "RESERVATION", "title": "예약 확정", "message": "오전 10:00 요가 수업 예약이 확정되었습니다.", "subMessage": "강사: 김수연 | 스튜디오 A", "isRead": false, "relatedUrl": "/reservations/42", "createdAt": "2026-02-17T13:55:00", "timeAgo": "방금 전"}], "totalElements": 15, "totalPages": 1, "hasNext": false}',
 "200 OK","빈 목록 → content: []","timeAgo 서버 계산"],

["NOTI-API-002","PATCH","/api/v1/notifications/{notificationId}/read","알림 읽음",
 "notificationId: long","-","-",
 "PATCH /api/v1/notifications/1/read",
 '{"notificationId": "long", "isRead": true}',
 '{"notificationId": 1, "isRead": true}',
 "200 OK\n404 Not Found","미존재 → 404","이미 읽음이면 200 (멱등)"],

["NOTI-API-003","PATCH","/api/v1/notifications/read-all","전체 읽음",
 "-","-","-",
 "PATCH /api/v1/notifications/read-all",
 '{"message": "모든 알림을 읽음 처리했습니다.", "updatedCount": "int"}',
 '{"message": "모든 알림을 읽음 처리했습니다.", "updatedCount": 5}',
 "200 OK","-","모두 읽음으로 표시 버튼"],

["NOTI-API-004","GET","/api/v1/notifications/unread-count","미읽음 수",
 "-","-","-",
 "GET /api/v1/notifications/unread-count",
 '{"unreadCount": "int"}',
 '{"unreadCount": 3}',
 "200 OK","-","벨 아이콘 배지 숫자"],

# ===== 공지사항 (ANN) 2개 =====
["ANN-API-001","GET","/api/v1/announcements","공지 목록",
 "-","page: int (기본 0)\nsize: int (기본 20)\ntag: NOTICE|EVENT (선택)",
 "-",
 "GET /api/v1/announcements?page=0&size=20",
 '{\n "content": [{\n  "announcementId": "long", "title": "string",\n  "tag": "NOTICE|EVENT",\n  "imageUrl": "string|null",\n  "isNew": "boolean",\n  "publishedAt": "datetime",\n  "summary": "string"\n }],\n "totalElements": "long", "totalPages": "int", "hasNext": "boolean"\n}',
 '{"content": [{"announcementId": 1, "title": "센터 정기 휴무 안내", "tag": "NOTICE", "imageUrl": "/img/notice1.jpg", "isNew": true, "publishedAt": "2023-12-10T09:00:00", "summary": "12월 15일 정기 휴무..."}], "totalElements": 5, "totalPages": 1, "hasNext": false}',
 "200 OK","-","공지사항 탭 카드 리스트"],

["ANN-API-002","GET","/api/v1/announcements/{announcementId}","공지 상세",
 "announcementId: long","-","-",
 "GET /api/v1/announcements/1",
 '{\n "announcementId": "long", "title": "string",\n "content": "string (HTML/Markdown)",\n "tag": "NOTICE|EVENT",\n "imageUrl": "string|null",\n "publishedAt": "datetime"\n}',
 '{"announcementId": 1, "title": "센터 정기 휴무 안내", "content": "12월 15일(금) 정기 휴무 및 대청소...", "tag": "NOTICE", "imageUrl": "/img/notice1.jpg", "publishedAt": "2023-12-10T09:00:00"}',
 "200 OK\n404 Not Found","미존재 → 404","공지 전체 내용"],

# ===== 문의 (INQ) 4개 =====
["INQ-API-001","GET","/api/v1/inquiries/me","내 문의 목록",
 "-","page: int (기본 0)\nsize: int (기본 20)\nperiod: int (선택, 월 수, 기본 3)\nstatus: RECEIVED|IN_PROGRESS|REPLIED (선택)",
 "-",
 "GET /api/v1/inquiries/me?page=0&period=3",
 '{\n "content": [{\n  "inquiryId": "long",\n  "category": "string",\n  "title": "string",\n  "status": "RECEIVED|IN_PROGRESS|REPLIED",\n  "contentPreview": "string",\n  "createdAt": "datetime"\n }],\n "totalElements": "long", "totalPages": "int", "hasNext": "boolean"\n}',
 '{"content": [{"inquiryId": 1, "category": "회원권", "title": "회원권 기간 연장 문의", "status": "REPLIED", "contentPreview": "3개월 정기권 연장 시...", "createdAt": "2024-05-12T10:00:00"}], "totalElements": 3, "totalPages": 1, "hasNext": false}',
 "200 OK","문의 없음 → content: []","최근 3개월 기본"],

["INQ-API-002","POST","/api/v1/inquiries","새 문의 작성",
 "-","-",
 '{\n "category": "string (필수, 시설|예약|회원권|기타)",\n "title": "string (필수)",\n "content": "string (필수)",\n "imageUrl": "string|null"\n}',
 '{"category": "회원권", "title": "회원권 기간 연장 문의드립니다.", "content": "3개월 정기권 연장 시 할인 혜택이 있나요?", "imageUrl": null}',
 '{"inquiryId": "long", "status": "RECEIVED", "createdAt": "datetime"}',
 '{"inquiryId": 5, "status": "RECEIVED", "createdAt": "2026-02-17T14:30:00"}',
 "201 Created\n400 Bad Request",
 "제목 누락 → 400\n내용 누락 → 400\n카테고리 미선택 → 400",
 "접수 완료 토스트 표시"],

["INQ-API-003","GET","/api/v1/inquiries/{inquiryId}","문의 상세",
 "inquiryId: long","-","-",
 "GET /api/v1/inquiries/1",
 '{\n "inquiryId": "long", "category": "string",\n "title": "string", "content": "string",\n "imageUrl": "string|null",\n "status": "string", "createdAt": "datetime",\n "reply": {\n  "replyId": "long|null",\n  "adminName": "string",\n  "content": "string",\n  "createdAt": "datetime"\n } | null\n}',
 '{"inquiryId": 1, "category": "회원권", "title": "회원권 기간 연장 문의", "content": "3개월 정기권 연장 시...", "imageUrl": null, "status": "REPLIED", "createdAt": "2024-05-12T10:00:00", "reply": {"replyId": 1, "adminName": "관리자", "content": "안녕하세요, 재등록 시 10% 할인...", "createdAt": "2024-05-13T09:00:00"}}',
 "200 OK\n403 Forbidden\n404 Not Found",
 "타인 문의 → 403\n미존재 → 404",
 "reply: null이면 미답변\nREPLIED면 reply 포함"],

["INQ-API-004","POST","/api/v1/inquiries/upload-image","문의 이미지 업로드",
 "-","-",
 "multipart/form-data\nfile: 이미지 파일 (필수, JPG/PNG/WEBP, 최대 5MB)",
 "Content-Type: multipart/form-data\n[file binary]",
 '{"imageUrl": "string"}',
 '{"imageUrl": "https://cdn.smithlife.co.kr/inquiries/img123.jpg"}',
 "200 OK\n413 Payload Too Large\n415 Unsupported",
 "5MB 초과 → 413\n미지원 형식 → 415",
 "문의 작성 전 이미지 선업로드\nURL을 INQ-API-002에 전달"],

# ===== FAQ 1개 =====
["FAQ-API-001","GET","/api/v1/faqs","FAQ 목록",
 "-","category: string (선택, 회원권|예약|시설|기타)",
 "-",
 "GET /api/v1/faqs?category=회원권",
 '{"faqs": [{"faqId": "long", "category": "string", "question": "string", "answer": "string", "orderIndex": "int"}]}',
 '{"faqs": [{"faqId": 1, "category": "회원권", "question": "회원권 일시정지는 어떻게 하나요?", "answer": "내정보 > 회원권 관리에서 일시정지를 신청...", "orderIndex": 1}]}',
 "200 OK","결과 없음 → faqs: []","아코디언 형태 표시"],

# ===== 챗봇 (CHAT) 2개 [신규] =====
["CHAT-API-001","POST","/api/v1/chatbot/sessions","챗봇 세션 생성",
 "-","-","-",
 "POST /api/v1/chatbot/sessions",
 '{"chatSessionId": "string (UUID)", "createdAt": "datetime", "welcomeMessage": "string"}',
 '{"chatSessionId": "550e8400-e29b-41d4-a716-446655440000", "createdAt": "2026-02-17T15:00:00", "welcomeMessage": "안녕하세요! SmithLife 챗봇입니다. 무엇을 도와드릴까요?"}',
 "201 Created","-","챗봇 상담 화면 진입 시 호출"],

["CHAT-API-002","POST","/api/v1/chatbot/sessions/{chatSessionId}/messages","챗봇 메시지",
 "chatSessionId: string (UUID)","-",
 '{"message": "string (필수)"}',
 '{"message": "회원권 연장은 어떻게 하나요?"}',
 '{\n "messageId": "string",\n "userMessage": "string",\n "botResponse": "string",\n "suggestions": ["string"],\n "createdAt": "datetime"\n}',
 '{"messageId": "msg-001", "userMessage": "회원권 연장은 어떻게 하나요?", "botResponse": "내정보 > 회원권 관리에서 연장하실 수 있습니다.", "suggestions": ["회원권 종류 알려줘", "가격이 궁금해요"], "createdAt": "2026-02-17T15:00:30"}',
 "200 OK\n400 Bad Request\n404 Not Found",
 "빈 메시지 → 400\n만료 세션 → 404",
 "suggestions: 후속 질문 추천\nAI/규칙 기반 응답"],

# ===== 디바이스/푸시 (DEV) 2개 [신규] =====
["DEV-API-001","POST","/api/v1/devices/fcm-token","FCM 토큰 등록",
 "-","-",
 '{\n "token": "string (필수)",\n "platform": "IOS|ANDROID (필수)",\n "deviceId": "string (필수)"\n}',
 '{"token": "fMC-abc123...", "platform": "ANDROID", "deviceId": "device-xyz-789"}',
 '{"message": "토큰이 등록되었습니다."}',
 '{"message": "토큰이 등록되었습니다."}',
 "200 OK\n400 Bad Request",
 "토큰 누락 → 400\n플랫폼 누락 → 400",
 "로그인 성공 후 호출\n기존 동일 deviceId 토큰 갱신"],

["DEV-API-002","DELETE","/api/v1/devices/fcm-token","FCM 토큰 삭제",
 "-","-",
 '{"deviceId": "string (필수)"}',
 '{"deviceId": "device-xyz-789"}',
 '{"message": "토큰이 삭제되었습니다."}',
 '{"message": "토큰이 삭제되었습니다."}',
 "200 OK","이미 삭제 → 200 (멱등)","로그아웃 시 호출\n해당 기기 푸시 중단"],

# ===== 앱 공통 (APP) 3개 [신규] =====
["APP-API-001","GET","/api/v1/app/version-check","앱 버전 체크",
 "-","platform: IOS|ANDROID (필수)\nversion: string (필수, 현재 앱 버전)",
 "-",
 "GET /api/v1/app/version-check?platform=ANDROID&version=1.2.4",
 '{\n "latestVersion": "string",\n "minVersion": "string",\n "forceUpdate": "boolean",\n "updateUrl": "string",\n "message": "string|null"\n}',
 '{"latestVersion": "1.3.0", "minVersion": "1.1.0", "forceUpdate": false, "updateUrl": "https://play.google.com/store/apps/details?id=...", "message": "새로운 기능이 추가되었습니다!"}',
 "200 OK","잘못된 platform → 400","forceUpdate=true면 업데이트 강제\n인증 불필요"],

["APP-API-002","GET","/api/v1/app/terms","이용약관/정책 조회",
 "-","type: service|privacy (필수)",
 "-",
 "GET /api/v1/app/terms?type=service",
 '{\n "type": "string",\n "title": "string",\n "content": "string (HTML/Markdown)",\n "version": "string",\n "updatedAt": "datetime"\n}',
 '{"type": "service", "title": "이용약관", "content": "<h1>제1조 목적</h1>...", "version": "2.0", "updatedAt": "2026-01-01T00:00:00"}',
 "200 OK\n400 Bad Request","잘못된 type → 400","인증 불필요\nHTML 렌더링"],

["APP-API-003","GET","/api/v1/app/terms/versions","약관 버전 목록",
 "-","-","-",
 "GET /api/v1/app/terms/versions",
 '{"versions": [{"type": "string", "version": "string", "title": "string", "updatedAt": "datetime"}]}',
 '{"versions": [{"type": "service", "version": "2.0", "title": "이용약관", "updatedAt": "2026-01-01T00:00:00"}, {"type": "privacy", "version": "1.5", "title": "개인정보처리방침", "updatedAt": "2025-12-01T00:00:00"}]}',
 "200 OK","-","약관 변경 이력 확인용"],
]

# ── 데이터 쓰기 ──
for r, row in enumerate(D, 2):
    for c, val in enumerate(row, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        cell.font = bf; cell.alignment = wa; cell.border = tb
        if r % 2 == 0: cell.fill = af
    # Method 스타일
    mc = ws2.cell(row=r, column=2)
    m = str(mc.value).upper() if mc.value else ""
    if m in MF: mc.font = MF[m]; mc.alignment = ca

# ── 저장 ──
output = r"c:\SSAFY\personalProject\SmithLife2\SmithLife_API명세서_v2.0.xlsx"
wb.save(output)
print(f"Part 2 완료: 전체 {len(D)}개 API 상세 명세 작성")
print(f"저장 위치: {output}")
