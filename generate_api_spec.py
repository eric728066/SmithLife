"""
SmithLife API 명세서 Excel 생성 스크립트
요구사항명세서 v3.0 + DDL 기반 전체 API 엔드포인트 도출
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 스타일 정의 ──
GOLD = "D4A04A"
DARK = "2C2C2C"
WHITE = "FFFFFF"
LIGHT_GOLD = "FDF6E3"
LIGHT_GRAY = "F5F5F5"
GREEN = "27AE60"
BLUE = "2980B9"
RED = "E74C3C"
ORANGE = "E67E22"

header_font = Font(name="맑은 고딕", bold=True, size=11, color=WHITE)
header_fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
sub_header_font = Font(name="맑은 고딕", bold=True, size=10, color=DARK)
sub_header_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
body_font = Font(name="맑은 고딕", size=10)
body_font_bold = Font(name="맑은 고딕", size=10, bold=True)
wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
gold_light_fill = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")

METHOD_FILLS = {
    "GET": PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid"),
    "POST": PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid"),
    "PUT": PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),
    "PATCH": PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),
    "DELETE": PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid"),
}
METHOD_FONTS = {
    "GET": Font(name="맑은 고딕", size=10, bold=True, color=BLUE),
    "POST": Font(name="맑은 고딕", size=10, bold=True, color=GREEN),
    "PUT": Font(name="맑은 고딕", size=10, bold=True, color=ORANGE),
    "PATCH": Font(name="맑은 고딕", size=10, bold=True, color=ORANGE),
    "DELETE": Font(name="맑은 고딕", size=10, bold=True, color=RED),
}


def style_sheet(ws, headers, col_widths):
    """공통 헤더 스타일 적용"""
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_rows(ws, data, start_row=2):
    """데이터 행 작성"""
    for r, row_data in enumerate(data, start_row):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = body_font
            cell.alignment = wrap_align
            cell.border = thin_border
            if r % 2 == 0:
                cell.fill = alt_fill
        # Method 열(2번째) 특수 스타일
        method_cell = ws.cell(row=r, column=2)
        method = str(method_cell.value).upper() if method_cell.value else ""
        if method in METHOD_FONTS:
            method_cell.font = METHOD_FONTS[method]
            method_cell.alignment = center_align


# ============================================================
# Sheet 1: API 총괄 목록
# ============================================================
ws1 = wb.active
ws1.title = "API 총괄 목록"

headers1 = [
    "No", "대분류", "중분류", "API ID", "HTTP Method", "엔드포인트 (URI)",
    "기능명", "설명", "인증 필요", "관련 요구사항 ID"
]
widths1 = [5, 10, 12, 14, 10, 38, 18, 50, 8, 20]

api_list = [
    # ── 인증 (AUTH) ──
    [1, "인증", "로그인", "AUTH-API-001", "POST", "/api/v1/auth/login",
     "로그인", "이메일/비밀번호로 로그인하여 JWT(Access+Refresh) 토큰 발급", "X", "AUTH-001~006"],
    [2, "인증", "세션", "AUTH-API-002", "POST", "/api/v1/auth/token/refresh",
     "토큰 갱신", "Refresh Token으로 새 Access Token 발급", "O (Refresh)", "SPL-003, NFR-007"],
    [3, "인증", "로그아웃", "AUTH-API-003", "POST", "/api/v1/auth/logout",
     "로그아웃", "서버 측 Refresh Token 무효화 처리", "O", "SET-007"],
    [4, "인증", "회원가입", "AUTH-API-004", "POST", "/api/v1/auth/signup",
     "회원가입", "이름/전화번호/이메일/비밀번호로 신규 회원 등록", "X", "AUTH-009~016"],
    [5, "인증", "회원가입", "AUTH-API-005", "GET", "/api/v1/auth/check-email?email={email}",
     "이메일 중복 확인", "가입 시 이메일 중복 여부 실시간 체크", "X", "AUTH-012, AUTH-015"],
    [6, "인증", "계정 찾기", "AUTH-API-006", "POST", "/api/v1/auth/find-email",
     "아이디(이메일) 찾기", "전화번호로 등록된 이메일 조회", "X", "AUTH-018"],
    [7, "인증", "계정 찾기", "AUTH-API-007", "POST", "/api/v1/auth/reset-password/request",
     "비밀번호 재설정 요청", "등록된 이메일로 재설정 링크/코드 발송", "X", "AUTH-019"],
    [8, "인증", "계정 찾기", "AUTH-API-008", "POST", "/api/v1/auth/reset-password/confirm",
     "비밀번호 재설정 확인", "인증 코드 확인 후 새 비밀번호 설정", "X", "AUTH-019"],

    # ── 홈 (HOME) ──
    [9, "홈", "오늘의 요약", "HOME-API-001", "GET", "/api/v1/home/summary",
     "홈 요약 조회", "다음 예약, 혼잡도, 회원권 남은 기간 등 홈 화면 통합 데이터 조회", "O", "HOME-003~006, HOME-010"],
    [10, "홈", "공지사항", "HOME-API-002", "GET", "/api/v1/home/announcements?limit={n}",
     "홈 공지 미리보기", "홈 가로 스크롤용 최신 공지 N건 조회", "O", "HOME-007, HOME-008"],

    # ── 사용자 (USER) ──
    [11, "사용자", "프로필", "USER-API-001", "GET", "/api/v1/users/me",
     "내 정보 조회", "로그인 사용자의 프로필, 회원권, 참석율, 오늘의 운동량 통합 조회", "O", "MY-001~003"],
    [12, "사용자", "프로필", "USER-API-002", "PUT", "/api/v1/users/me",
     "내 정보 수정", "이름, 전화번호, 프로필 이미지 등 프로필 정보 수정", "O", "MY-007, SET-002"],
    [13, "사용자", "프로필", "USER-API-003", "POST", "/api/v1/users/me/profile-image",
     "프로필 이미지 업로드", "프로필 사진 파일 업로드 (multipart/form-data)", "O", "MY-007"],
    [14, "사용자", "계정", "USER-API-004", "PUT", "/api/v1/users/me/password",
     "비밀번호 변경", "현재 비밀번호 확인 후 새 비밀번호로 변경", "O", "SET-006"],
    [15, "사용자", "계정", "USER-API-005", "DELETE", "/api/v1/users/me",
     "회원 탈퇴", "비밀번호 재확인 후 계정 비활성화(soft delete) 처리", "O", "SET-008"],

    # ── 회원권 (MEMBERSHIP) ──
    [16, "회원권", "조회", "MBSP-API-001", "GET", "/api/v1/memberships/me",
     "내 회원권 조회", "현재 활성 회원권 정보 (종류, 시작일, 만료일, D-day) 조회", "O", "HOME-006, MY-001"],
    [17, "회원권", "관리", "MBSP-API-002", "POST", "/api/v1/memberships",
     "회원권 등록/연장", "신규 회원권 등록 또는 기존 회원권 연장 처리", "O", "NOTI-004"],
    [18, "회원권", "관리", "MBSP-API-003", "PATCH", "/api/v1/memberships/{membershipId}/pause",
     "회원권 일시정지", "회원권 상태를 PAUSED로 변경 (최대 30일)", "O", "FAQ"],

    # ── 출석/QR (ATTENDANCE) ──
    [19, "출석", "QR 체크인", "ATT-API-001", "POST", "/api/v1/attendance/qr-token",
     "QR 토큰 생성", "사용자 고유 1회성 QR 코드 토큰 생성 (30초~1분 유효)", "O", "QR-003, QR-004, NFR-009"],
    [20, "출석", "QR 체크인", "ATT-API-002", "POST", "/api/v1/attendance/check-in",
     "QR 체크인 처리", "QR 스캔 후 출석 체크인 기록 생성 (키오스크/직원용)", "O (Admin)", "QR-005"],
    [21, "출석", "QR 체크인", "ATT-API-003", "PATCH", "/api/v1/attendance/{attendanceId}/check-out",
     "체크아웃 처리", "퇴장 시 체크아웃 시간 기록", "O", "QR-005"],
    [22, "출석", "조회", "ATT-API-004", "GET", "/api/v1/attendance/me?month={yyyy-MM}",
     "내 출석 기록 조회", "월별 출석 기록 리스트 조회 (참석율 계산용)", "O", "MY-002"],
    [23, "출석", "조회", "ATT-API-005", "GET", "/api/v1/attendance/me/status",
     "현재 출석 상태 조회", "현재 체크인 여부 및 진행 중인 출석 정보 조회", "O", "QR-005"],

    # ── 시설 (FACILITY) ──
    [24, "시설", "조회", "FAC-API-001", "GET", "/api/v1/facilities",
     "시설 목록 조회", "전체 시설(메인 피트니스 존, 스튜디오 A/B 등) 목록 조회", "O", "HOME-004, RESV-001"],
    [25, "시설", "혼잡도", "FAC-API-002", "GET", "/api/v1/facilities/{facilityId}/congestion",
     "시설 혼잡도 조회", "특정 시설의 현재 혼잡도(원활/보통/혼잡) 실시간 조회", "O", "HOME-005, RESV-003"],

    # ── 시간 슬롯 (TIME SLOT) ──
    [26, "예약", "시간표", "SLOT-API-001", "GET", "/api/v1/facilities/{facilityId}/slots?date={yyyy-MM-dd}",
     "시간 슬롯 조회", "특정 시설/날짜의 시간 슬롯 목록 + 혼잡도(현재인원/최대인원) 조회", "O", "RESV-002, RESV-003, RESV-006"],

    # ── 예약 (RESERVATION) ──
    [27, "예약", "예약", "RESV-API-001", "POST", "/api/v1/reservations",
     "예약 생성", "선택한 시간 슬롯(복수 가능)으로 예약 생성, 예약번호 발급", "O", "RESV-009~010, RESV-012"],
    [28, "예약", "예약", "RESV-API-002", "GET", "/api/v1/reservations/me?date={yyyy-MM-dd}",
     "내 예약 목록 조회", "특정 날짜 기준 내 예약 리스트 조회 (취소 모달용 포함)", "O", "RESV-005, RESV-014, HOME-004"],
    [29, "예약", "예약", "RESV-API-003", "GET", "/api/v1/reservations/me/upcoming",
     "다음 예약 조회", "가장 가까운 미래 예약 1건 조회 (홈 카드용)", "O", "HOME-004"],
    [30, "예약", "취소", "RESV-API-004", "PATCH", "/api/v1/reservations/{reservationId}/cancel",
     "예약 취소", "예약 상태를 CANCELLED로 변경, 슬롯 인원 차감", "O", "RESV-015~017"],
    [31, "예약", "상세", "RESV-API-005", "GET", "/api/v1/reservations/{reservationId}",
     "예약 상세 조회", "예약번호, 시설명, 날짜, 시간, 상태 등 상세 정보 조회", "O", "NOTI-006"],

    # ── 이용내역 (USAGE HISTORY) ──
    [32, "이용내역", "조회", "USAGE-API-001", "GET", "/api/v1/usage-history/me?page={p}&size={s}",
     "이용내역 목록 조회", "최신순 페이지네이션 (무한스크롤) 이용내역 카드 리스트 조회", "O", "USAGE-001~006, HOME-009"],
    [33, "이용내역", "상세", "USAGE-API-002", "GET", "/api/v1/usage-history/{usageId}",
     "이용내역 상세 조회", "개별 이용내역 상세 정보 (시설명, 시간, 예약번호, 상태) 조회", "O", "USAGE-002"],

    # ── 운동 종목 (EXERCISE) ──
    [34, "운동", "종목", "EXER-API-001", "GET", "/api/v1/exercises?bodyPart={part}&keyword={kw}",
     "운동 종목 검색", "부위/키워드 필터로 운동 종목 목록 검색 (운동 추가 시 사용)", "O", "WORK-009, RTNC-005"],
    [35, "운동", "종목", "EXER-API-002", "GET", "/api/v1/exercises/{exerciseId}",
     "운동 종목 상세", "운동명, 부위, 장비, 이미지, 설명 등 상세 정보 조회", "O", "EXER-001"],

    # ── 루틴 (ROUTINE) ──
    [36, "운동", "추천 루틴", "RTN-API-001", "GET", "/api/v1/routines/recommended",
     "추천 루틴 목록", "시스템 추천 루틴 목록 조회 (2열 그리드용)", "O", "RTN-001~005, WORK-004~005"],
    [37, "운동", "루틴 검색", "RTN-API-002", "GET", "/api/v1/routines?keyword={kw}&goal={goal}",
     "루틴 검색", "키워드/목표 필터로 루틴 검색 (공개 루틴 포함)", "O", "RTN-002"],
    [38, "운동", "루틴 상세", "RTN-API-003", "GET", "/api/v1/routines/{routineId}",
     "루틴 상세 조회", "루틴 기본정보 + 운동 구성 리스트 + 찜 여부 통합 조회", "O", "RTND-001~005"],
    [39, "운동", "루틴 생성", "RTN-API-004", "POST", "/api/v1/routines",
     "나만의 루틴 생성", "루틴명/목표/운동목록/공유여부 포함 커스텀 루틴 저장", "O", "RTNC-001~008"],
    [40, "운동", "루틴 관리", "RTN-API-005", "PUT", "/api/v1/routines/{routineId}",
     "루틴 수정", "내가 만든 루틴의 이름/목표/운동구성/공유 설정 수정", "O", "RTNC-006"],
    [41, "운동", "루틴 관리", "RTN-API-006", "DELETE", "/api/v1/routines/{routineId}",
     "루틴 삭제", "내가 만든 루틴 삭제", "O", "RTNC-001"],
    [42, "운동", "나의 루틴", "RTN-API-007", "GET", "/api/v1/routines/me",
     "내 루틴 목록", "내가 생성한 루틴 목록 조회", "O", "RTN-006"],

    # ── 루틴 찜 (FAVORITE) ──
    [43, "운동", "찜", "FAV-API-001", "POST", "/api/v1/routines/{routineId}/favorite",
     "루틴 찜 추가", "루틴을 즐겨찾기에 추가 (하트 채우기)", "O", "RTND-006"],
    [44, "운동", "찜", "FAV-API-002", "DELETE", "/api/v1/routines/{routineId}/favorite",
     "루틴 찜 해제", "루틴을 즐겨찾기에서 제거 (하트 비우기)", "O", "RTND-006"],
    [45, "운동", "찜", "FAV-API-003", "GET", "/api/v1/routines/favorites",
     "찜한 루틴 목록", "내가 찜한 루틴 전체 목록 조회", "O", "RTND-006"],

    # ── 운동 세션 (WORKOUT SESSION) ──
    [46, "운동", "세션", "SESS-API-001", "POST", "/api/v1/workout-sessions",
     "운동 세션 시작", "새 운동 세션 생성 (루틴 기반 or 빈 세션). 루틴ID 전달 시 운동 자동 구성", "O", "RTND-007, WORK-011"],
    [47, "운동", "세션", "SESS-API-002", "GET", "/api/v1/workout-sessions/active",
     "활성 세션 조회", "현재 진행 중인(ACTIVE/PAUSED) 세션 정보 + 운동 리스트 조회", "O", "WORK-001~003, WORK-006"],
    [48, "운동", "세션", "SESS-API-003", "PATCH", "/api/v1/workout-sessions/{sessionId}/end",
     "운동 세션 종료", "세션 상태를 COMPLETED로 변경, 총 시간/볼륨/칼로리 최종 계산", "O", "WORK-010"],
    [49, "운동", "세션", "SESS-API-004", "PATCH", "/api/v1/workout-sessions/{sessionId}/pause",
     "운동 세션 일시정지", "세션 상태를 PAUSED로 변경", "O", "WORK-001"],
    [50, "운동", "세션", "SESS-API-005", "PATCH", "/api/v1/workout-sessions/{sessionId}/resume",
     "운동 세션 재개", "세션 상태를 ACTIVE로 재변경", "O", "WORK-001"],

    # ── 세션 내 운동 (SESSION EXERCISE) ──
    [51, "운동", "세션 운동", "SE-API-001", "POST", "/api/v1/workout-sessions/{sessionId}/exercises",
     "세션에 운동 추가", "진행 중 세션에 새 운동 추가 (+ 운동 추가하기)", "O", "WORK-009"],
    [52, "운동", "세션 운동", "SE-API-002", "GET", "/api/v1/workout-sessions/{sessionId}/exercises",
     "세션 운동 목록 조회", "세션 내 운동 리스트 (완료/진행중/대기 상태 포함) 조회", "O", "WORK-006~008"],
    [53, "운동", "세션 운동", "SE-API-003", "PATCH", "/api/v1/session-exercises/{sessionExerciseId}/complete",
     "세션 운동 완료 처리", "특정 운동을 COMPLETED 상태로 변경", "O", "EXER-008"],
    [54, "운동", "세션 운동", "SE-API-004", "PATCH", "/api/v1/session-exercises/{sessionExerciseId}/start",
     "세션 운동 시작", "특정 운동을 IN_PROGRESS(UP NEXT → 진행중)로 변경", "O", "WORK-008"],
    [55, "운동", "세션 운동", "SE-API-005", "DELETE", "/api/v1/session-exercises/{sessionExerciseId}",
     "세션 운동 삭제", "세션에서 특정 운동 제거", "O", "WORK-009"],

    # ── 세트 기록 (EXERCISE SET) ──
    [56, "운동", "세트", "SET-API-001", "POST", "/api/v1/session-exercises/{sessionExerciseId}/sets",
     "세트 추가", "운동에 새 세트 행 추가 (+ 세트 추가하기)", "O", "EXER-005"],
    [57, "운동", "세트", "SET-API-002", "GET", "/api/v1/session-exercises/{sessionExerciseId}/sets",
     "세트 목록 조회", "특정 운동의 전체 세트(번호/KG/회/완료여부) 리스트 조회", "O", "EXER-002"],
    [58, "운동", "세트", "SET-API-003", "PATCH", "/api/v1/exercise-sets/{setId}",
     "세트 정보 수정", "KG, 횟수, 휴식시간 등 세트 데이터 수정", "O", "EXER-003"],
    [59, "운동", "세트", "SET-API-004", "PATCH", "/api/v1/exercise-sets/{setId}/complete",
     "세트 완료 체크", "세트 완료 처리 (초록 체크) + 완료 시각 기록 + 휴식 타이머 트리거", "O", "EXER-004, EXER-009"],
    [60, "운동", "세트", "SET-API-005", "DELETE", "/api/v1/exercise-sets/{setId}",
     "세트 삭제", "특정 세트 행 삭제", "O", "EXER-005"],

    # ── 개인 기록 (PERSONAL RECORD) ──
    [61, "운동", "개인기록", "PR-API-001", "GET", "/api/v1/personal-records?exerciseId={id}",
     "운동별 최고 기록 조회", "특정 운동의 1RM/최대볼륨/최대횟수 등 최고 기록 조회", "O", "EXER-006"],
    [62, "운동", "개인기록", "PR-API-002", "GET", "/api/v1/personal-records/me",
     "전체 개인 기록 목록", "내 모든 운동의 개인 최고 기록 리스트 조회", "O", "EXER-006"],

    # ── 운동 리포트 (WORKOUT REPORT) ──
    [63, "리포트", "요약", "RPT-API-001", "GET", "/api/v1/workout-reports/{sessionId}",
     "운동 리포트 조회", "세션 완료 후 리포트 (시간/볼륨/칼로리/운동요약) 조회", "O", "RPT-001~004, RPT-006"],
    [64, "리포트", "차트", "RPT-API-002", "GET", "/api/v1/workout-reports/weekly?date={yyyy-MM-dd}",
     "주간 운동량 차트 데이터", "월~일 요일별 운동량 + 지난주 대비 변화율(%) 조회", "O", "RPT-003"],
    [65, "리포트", "공유", "RPT-API-003", "POST", "/api/v1/workout-reports/{reportId}/share",
     "리포트 공유 이미지 생성", "리포트를 이미지로 캡처/생성하여 공유 URL 반환", "O", "RPT-005"],
    [66, "리포트", "오늘", "RPT-API-004", "GET", "/api/v1/workout-reports/today",
     "오늘의 운동량 조회", "오늘 총 볼륨(kg) 등 내정보 카드용 데이터 조회", "O", "MY-003"],

    # ── 설정 (SETTINGS) ──
    [67, "설정", "사용자 설정", "SETT-API-001", "GET", "/api/v1/settings",
     "설정 조회", "알림/다크모드/언어 등 현재 설정값 조회", "O", "SET-003~005"],
    [68, "설정", "사용자 설정", "SETT-API-002", "PATCH", "/api/v1/settings",
     "설정 변경", "알림 ON/OFF, 다크모드(SYSTEM/ON/OFF), 언어(ko/en/ja) 변경", "O", "SET-003~005, SET-012"],

    # ── 알림 (NOTIFICATION) ──
    [69, "알림", "개인 알림", "NOTI-API-001", "GET", "/api/v1/notifications?page={p}&size={s}",
     "알림 목록 조회", "개인 알림 페이지네이션 목록 (읽음/미읽음 구분 포함) 조회", "O", "NOTI-001~005"],
    [70, "알림", "개인 알림", "NOTI-API-002", "PATCH", "/api/v1/notifications/{notificationId}/read",
     "알림 읽음 처리", "개별 알림을 읽음 상태로 변경", "O", "NOTI-005"],
    [71, "알림", "개인 알림", "NOTI-API-003", "PATCH", "/api/v1/notifications/read-all",
     "전체 알림 읽음 처리", "모든 미읽음 알림을 일괄 읽음 처리 (모두 읽음으로 표시)", "O", "NOTI-002"],
    [72, "알림", "개인 알림", "NOTI-API-004", "GET", "/api/v1/notifications/unread-count",
     "미읽음 알림 수 조회", "알림 벨 아이콘 배지 숫자용 미읽음 개수 조회", "O", "HOME-002"],

    # ── 공지사항 (ANNOUNCEMENT) ──
    [73, "공지사항", "목록", "ANN-API-001", "GET", "/api/v1/announcements?page={p}&size={s}",
     "공지사항 목록 조회", "공지사항 탭 카드 리스트 (태그/이미지/날짜 포함) 페이지네이션 조회", "O", "NOTI-007, SET-009"],
    [74, "공지사항", "상세", "ANN-API-002", "GET", "/api/v1/announcements/{announcementId}",
     "공지사항 상세 조회", "공지 전체 내용 (제목/내용/이미지/게시일) 조회", "O", "NOTI-008, HOME-008"],

    # ── 1:1 문의 (INQUIRY) ──
    [75, "문의", "1:1 문의", "INQ-API-001", "GET", "/api/v1/inquiries/me?page={p}&size={s}&period={m}",
     "내 문의 목록 조회", "기간 필터 적용 문의 내역 카드 리스트 (상태태그 포함) 페이지네이션 조회", "O", "INQ-004~006"],
    [76, "문의", "1:1 문의", "INQ-API-002", "POST", "/api/v1/inquiries",
     "새 문의 작성", "제목/내용/카테고리/이미지 첨부로 새 1:1 문의 접수", "O", "INQ-007~009"],
    [77, "문의", "1:1 문의", "INQ-API-003", "GET", "/api/v1/inquiries/{inquiryId}",
     "문의 상세 조회", "문의 내용 + 답변 내용 (있으면) 통합 조회", "O", "INQ-006"],
    [78, "문의", "이미지", "INQ-API-004", "POST", "/api/v1/inquiries/upload-image",
     "문의 이미지 업로드", "문의 첨부 이미지 파일 업로드 (multipart/form-data)", "O", "INQ-008"],

    # ── FAQ ──
    [79, "문의", "FAQ", "FAQ-API-001", "GET", "/api/v1/faqs",
     "FAQ 목록 조회", "카테고리별 자주 묻는 질문 전체 목록 (아코디언용) 조회", "O", "INQ-002"],
]

style_sheet(ws1, headers1, widths1)
write_rows(ws1, api_list)

# ============================================================
# Sheet 2: API 상세 명세서
# ============================================================
ws2 = wb.create_sheet("API 상세 명세")
headers2 = [
    "API ID", "Method", "엔드포인트", "기능명",
    "Path Parameters", "Query Parameters",
    "Request Body (JSON)", "Request 예시",
    "Response Body (JSON)", "Response 예시",
    "상태코드", "에러 케이스", "비고"
]
widths2 = [13, 8, 35, 14, 20, 22, 40, 40, 40, 45, 18, 30, 22]

detail_data = [
    # AUTH-API-001: 로그인
    ["AUTH-API-001", "POST", "/api/v1/auth/login", "로그인",
     "-", "-",
     '{\n  "email": "string (필수)",\n  "password": "string (필수)"\n}',
     '{\n  "email": "smith@example.com",\n  "password": "mypass123"\n}',
     '{\n  "accessToken": "string (JWT)",\n  "refreshToken": "string",\n  "tokenType": "Bearer",\n  "expiresIn": 3600,\n  "user": {\n    "userId": "long",\n    "name": "string",\n    "email": "string",\n    "profileImageUrl": "string|null",\n    "role": "USER|ADMIN"\n  }\n}',
     '{\n  "accessToken": "eyJhbG...",\n  "refreshToken": "dGhpcyBp...",\n  "tokenType": "Bearer",\n  "expiresIn": 3600,\n  "user": {\n    "userId": 1,\n    "name": "김스미스",\n    "email": "smith@example.com",\n    "profileImageUrl": null,\n    "role": "USER"\n  }\n}',
     "200 OK\n401 Unauthorized\n400 Bad Request\n500 Internal Server Error",
     "이메일 미존재 → 401\n비밀번호 불일치 → 401\n이메일 형식 오류 → 400\n비활성 계정 → 403",
     "중복 클릭 방지 필요\nRefresh Token httpOnly cookie 권장"],

    # AUTH-API-002: 토큰 갱신
    ["AUTH-API-002", "POST", "/api/v1/auth/token/refresh", "토큰 갱신",
     "-", "-",
     '{\n  "refreshToken": "string (필수)"\n}',
     '{\n  "refreshToken": "dGhpcyBp..."\n}',
     '{\n  "accessToken": "string (JWT)",\n  "refreshToken": "string",\n  "expiresIn": 3600\n}',
     '{\n  "accessToken": "eyJhbG..new..",\n  "refreshToken": "newRefresh...",\n  "expiresIn": 3600\n}',
     "200 OK\n401 Unauthorized",
     "Refresh Token 만료 → 401\nRefresh Token 무효 → 401",
     "Refresh Token Rotation 적용"],

    # AUTH-API-003: 로그아웃
    ["AUTH-API-003", "POST", "/api/v1/auth/logout", "로그아웃",
     "-", "-",
     '{\n  "refreshToken": "string (필수)"\n}',
     '{\n  "refreshToken": "dGhpcyBp..."\n}',
     '{\n  "message": "로그아웃 완료"\n}',
     '{\n  "message": "로그아웃 완료"\n}',
     "200 OK\n401 Unauthorized",
     "이미 만료된 토큰 → 200 (멱등성)",
     "서버 측 Refresh Token 블랙리스트 처리"],

    # AUTH-API-004: 회원가입
    ["AUTH-API-004", "POST", "/api/v1/auth/signup", "회원가입",
     "-", "-",
     '{\n  "name": "string (필수)",\n  "phone": "string (필수, 010-XXXX-XXXX)",\n  "email": "string (필수, 이메일형식)",\n  "password": "string (필수, 6자 이상)"\n}',
     '{\n  "name": "김스미스",\n  "phone": "010-1234-5678",\n  "email": "smith@example.com",\n  "password": "mypass123"\n}',
     '{\n  "userId": "long",\n  "name": "string",\n  "email": "string",\n  "createdAt": "datetime"\n}',
     '{\n  "userId": 1,\n  "name": "김스미스",\n  "email": "smith@example.com",\n  "createdAt": "2026-02-17T10:00:00"\n}',
     "201 Created\n400 Bad Request\n409 Conflict",
     "이메일 중복 → 409\n전화번호 중복 → 409\n비밀번호 6자 미만 → 400\n필수값 누락 → 400",
     "비밀번호 bcrypt 해싱 저장\nUserSettings 자동 생성"],

    # AUTH-API-005: 이메일 중복 확인
    ["AUTH-API-005", "GET", "/api/v1/auth/check-email", "이메일 중복 확인",
     "-", "email: string (필수) - 확인할 이메일 주소",
     "-",
     "GET /api/v1/auth/check-email?email=smith@example.com",
     '{\n  "available": "boolean",\n  "message": "string"\n}',
     '{\n  "available": true,\n  "message": "사용 가능한 이메일입니다."\n}',
     "200 OK\n400 Bad Request",
     "이메일 형식 오류 → 400",
     "실시간 유효성 검증용 (디바운스 권장)"],

    # AUTH-API-006: 아이디 찾기
    ["AUTH-API-006", "POST", "/api/v1/auth/find-email", "아이디 찾기",
     "-", "-",
     '{\n  "phone": "string (필수)"\n}',
     '{\n  "phone": "010-1234-5678"\n}',
     '{\n  "email": "string (마스킹)",\n  "createdAt": "date"\n}',
     '{\n  "email": "sm***@example.com",\n  "createdAt": "2026-01-15"\n}',
     "200 OK\n404 Not Found",
     "등록되지 않은 전화번호 → 404",
     "이메일 부분 마스킹 처리"],

    # AUTH-API-007: 비밀번호 재설정 요청
    ["AUTH-API-007", "POST", "/api/v1/auth/reset-password/request", "비밀번호 재설정 요청",
     "-", "-",
     '{\n  "email": "string (필수)"\n}',
     '{\n  "email": "smith@example.com"\n}',
     '{\n  "message": "string",\n  "expiresIn": 300\n}',
     '{\n  "message": "재설정 코드가 이메일로 발송되었습니다.",\n  "expiresIn": 300\n}',
     "200 OK\n404 Not Found",
     "미등록 이메일 → 404\n연속 요청 제한 → 429",
     "5분 유효 코드 발송\nRate Limit 적용"],

    # AUTH-API-008: 비밀번호 재설정 확인
    ["AUTH-API-008", "POST", "/api/v1/auth/reset-password/confirm", "비밀번호 재설정 확인",
     "-", "-",
     '{\n  "email": "string (필수)",\n  "code": "string (필수)",\n  "newPassword": "string (필수, 6자 이상)"\n}',
     '{\n  "email": "smith@example.com",\n  "code": "A3F82K",\n  "newPassword": "newpass456"\n}',
     '{\n  "message": "비밀번호가 변경되었습니다."\n}',
     '{\n  "message": "비밀번호가 변경되었습니다."\n}',
     "200 OK\n400 Bad Request\n401 Unauthorized",
     "잘못된 코드 → 401\n만료된 코드 → 401\n비밀번호 규칙 미충족 → 400",
     "코드 사용 후 즉시 무효화"],

    # HOME-API-001: 홈 요약
    ["HOME-API-001", "GET", "/api/v1/home/summary", "홈 요약 조회",
     "-", "-",
     "-",
     "GET /api/v1/home/summary\nAuthorization: Bearer eyJhbG...",
     '{\n  "nextReservation": {\n    "reservationId": "long|null",\n    "startTime": "HH:mm",\n    "endTime": "HH:mm",\n    "facilityName": "string"\n  },\n  "congestion": {\n    "status": "LOW|MEDIUM|HIGH",\n    "label": "string"\n  },\n  "membership": {\n    "type": "string",\n    "dDay": "int",\n    "status": "ACTIVE|EXPIRED"\n  }\n}',
     '{\n  "nextReservation": {\n    "reservationId": 42,\n    "startTime": "09:00",\n    "endTime": "10:30",\n    "facilityName": "메인 피트니스 존"\n  },\n  "congestion": {\n    "status": "LOW",\n    "label": "원활합니다"\n  },\n  "membership": {\n    "type": "3개월 정기권",\n    "dDay": 45,\n    "status": "ACTIVE"\n  }\n}',
     "200 OK\n401 Unauthorized",
     "예약 없음 → nextReservation null\n회원권 없음 → membership null",
     "Pull-to-Refresh 대응\n혼잡도 실시간 반영"],

    # HOME-API-002: 홈 공지
    ["HOME-API-002", "GET", "/api/v1/home/announcements", "홈 공지 미리보기",
     "-", "limit: int (기본값 5) - 조회 건수",
     "-",
     "GET /api/v1/home/announcements?limit=5",
     '{\n  "announcements": [\n    {\n      "announcementId": "long",\n      "tag": "NOTICE|EVENT",\n      "title": "string",\n      "summary": "string (1~2줄)",\n      "publishedAt": "datetime"\n    }\n  ]\n}',
     '{\n  "announcements": [\n    {\n      "announcementId": 1,\n      "tag": "NOTICE",\n      "title": "설날 연휴 운영 시간 안내",\n      "summary": "2월 8일~10일 운영시간이...",\n      "publishedAt": "2026-02-07T09:00:00"\n    }\n  ]\n}',
     "200 OK",
     "-",
     "가로 스크롤 카드용"],

    # RESV-API-001: 예약 생성
    ["RESV-API-001", "POST", "/api/v1/reservations", "예약 생성",
     "-", "-",
     '{\n  "slotIds": ["long[] (필수, 1개 이상)"]\n}',
     '{\n  "slotIds": [101, 102]\n}',
     '{\n  "reservations": [\n    {\n      "reservationId": "long",\n      "reservationNo": "string",\n      "slotId": "long",\n      "date": "date",\n      "startTime": "HH:mm",\n      "endTime": "HH:mm",\n      "facilityName": "string",\n      "status": "CONFIRMED"\n    }\n  ]\n}',
     '{\n  "reservations": [\n    {\n      "reservationId": 42,\n      "reservationNo": "#8812",\n      "slotId": 101,\n      "date": "2026-02-10",\n      "startTime": "09:00",\n      "endTime": "10:00",\n      "facilityName": "메인 피트니스 존",\n      "status": "CONFIRMED"\n    }\n  ]\n}',
     "201 Created\n400 Bad Request\n409 Conflict",
     "만석 슬롯 → 409\n중복 예약 → 409\n빈 slotIds → 400",
     "복수 시간 동시 예약 가능\nslot current_count 증가"],

    # RESV-API-004: 예약 취소
    ["RESV-API-004", "PATCH", "/api/v1/reservations/{reservationId}/cancel", "예약 취소",
     "reservationId: long (필수)", "-",
     "-",
     "PATCH /api/v1/reservations/42/cancel",
     '{\n  "reservationId": "long",\n  "reservationNo": "string",\n  "status": "CANCELLED",\n  "cancelledAt": "datetime"\n}',
     '{\n  "reservationId": 42,\n  "reservationNo": "#8812",\n  "status": "CANCELLED",\n  "cancelledAt": "2026-02-10T08:30:00"\n}',
     "200 OK\n404 Not Found\n400 Bad Request",
     "이미 취소된 예약 → 400\n다른 사용자 예약 → 403\n완료된 예약 → 400",
     "취소 전 확인 다이얼로그 필수\nslot current_count 감소"],

    # SLOT-API-001: 시간 슬롯 조회
    ["SLOT-API-001", "GET", "/api/v1/facilities/{facilityId}/slots", "시간 슬롯 조회",
     "facilityId: long (필수)", "date: string (필수, yyyy-MM-dd)",
     "-",
     "GET /api/v1/facilities/1/slots?date=2026-02-10",
     '{\n  "facilityName": "string",\n  "date": "date",\n  "slots": [\n    {\n      "slotId": "long",\n      "startTime": "HH:mm",\n      "endTime": "HH:mm",\n      "maxCapacity": "int",\n      "currentCount": "int",\n      "congestion": "LOW|MEDIUM|HIGH",\n      "myReserved": "boolean"\n    }\n  ]\n}',
     '{\n  "facilityName": "메인 피트니스 존",\n  "date": "2026-02-10",\n  "slots": [\n    {\n      "slotId": 101,\n      "startTime": "09:00",\n      "endTime": "10:00",\n      "maxCapacity": 50,\n      "currentCount": 12,\n      "congestion": "LOW",\n      "myReserved": false\n    },\n    {\n      "slotId": 102,\n      "startTime": "10:00",\n      "endTime": "11:00",\n      "maxCapacity": 50,\n      "currentCount": 35,\n      "congestion": "MEDIUM",\n      "myReserved": true\n    }\n  ]\n}',
     "200 OK\n404 Not Found",
     "존재하지 않는 시설 → 404",
     "congestion: LOW(~40%), MEDIUM(40~80%), HIGH(80%~)\nmyReserved로 취소 가능 여부 판단"],

    # SESS-API-001: 운동 세션 시작
    ["SESS-API-001", "POST", "/api/v1/workout-sessions", "운동 세션 시작",
     "-", "-",
     '{\n  "routineId": "long|null (루틴 기반이면 전달)",\n  "sessionName": "string (필수)"\n}',
     '{\n  "routineId": 1,\n  "sessionName": "가슴 & 삼두 데이"\n}',
     '{\n  "sessionId": "long",\n  "sessionName": "string",\n  "status": "ACTIVE",\n  "startTime": "datetime",\n  "exercises": [\n    {\n      "sessionExerciseId": "long",\n      "exerciseId": "long",\n      "exerciseName": "string",\n      "bodyPart": "string",\n      "orderIndex": "int",\n      "targetSets": "int",\n      "targetReps": "int",\n      "status": "PENDING"\n    }\n  ]\n}',
     '{\n  "sessionId": 10,\n  "sessionName": "가슴 & 삼두 데이",\n  "status": "ACTIVE",\n  "startTime": "2026-02-17T14:00:00",\n  "exercises": [\n    {\n      "sessionExerciseId": 1,\n      "exerciseId": 1,\n      "exerciseName": "벤치 프레스",\n      "bodyPart": "가슴",\n      "orderIndex": 1,\n      "targetSets": 3,\n      "targetReps": 12,\n      "status": "PENDING"\n    }\n  ]\n}',
     "201 Created\n400 Bad Request\n409 Conflict",
     "이미 활성 세션 존재 → 409\n세션명 미입력 → 400",
     "routineId 전달 시 루틴의 운동 자동 복사\n한 번에 하나의 ACTIVE 세션만 허용"],

    # ATT-API-001: QR 토큰 생성
    ["ATT-API-001", "POST", "/api/v1/attendance/qr-token", "QR 토큰 생성",
     "-", "-",
     "-",
     "POST /api/v1/attendance/qr-token\nAuthorization: Bearer eyJhbG...",
     '{\n  "qrToken": "string",\n  "qrImageUrl": "string (base64 or URL)",\n  "expiresAt": "datetime",\n  "ttlSeconds": 60\n}',
     '{\n  "qrToken": "SL-QR-a3f82k9d...",\n  "qrImageUrl": "data:image/png;base64,...",\n  "expiresAt": "2026-02-17T14:01:00",\n  "ttlSeconds": 60\n}',
     "200 OK\n401 Unauthorized\n403 Forbidden",
     "회원권 만료 → 403\n비활성 회원 → 403",
     "1회성 + 60초 유효\n30초~1분 자동 갱신"],

    # NOTI-API-001: 알림 목록
    ["NOTI-API-001", "GET", "/api/v1/notifications", "알림 목록 조회",
     "-", "page: int (기본 0)\nsize: int (기본 20)",
     "-",
     "GET /api/v1/notifications?page=0&size=20",
     '{\n  "content": [\n    {\n      "notificationId": "long",\n      "type": "RESERVATION|MEMBERSHIP|WORKOUT|SYSTEM",\n      "title": "string",\n      "message": "string",\n      "subMessage": "string|null",\n      "isRead": "boolean",\n      "relatedUrl": "string|null",\n      "createdAt": "datetime",\n      "timeAgo": "string"\n    }\n  ],\n  "totalElements": "long",\n  "totalPages": "int",\n  "hasNext": "boolean"\n}',
     '{\n  "content": [\n    {\n      "notificationId": 1,\n      "type": "RESERVATION",\n      "title": "예약 확정",\n      "message": "오전 10:00 요가 수업 예약이 확정되었습니다.",\n      "subMessage": "강사: 김수연 | 스튜디오 A",\n      "isRead": false,\n      "relatedUrl": "/reservations/42",\n      "createdAt": "2026-02-17T13:55:00",\n      "timeAgo": "방금 전"\n    }\n  ],\n  "totalElements": 15,\n  "totalPages": 1,\n  "hasNext": false\n}',
     "200 OK",
     "-",
     "timeAgo 서버 계산 (방금전/N분전/N시간전/날짜)\nFCM/APNs 푸시 연동"],

    # INQ-API-002: 새 문의 작성
    ["INQ-API-002", "POST", "/api/v1/inquiries", "새 문의 작성",
     "-", "-",
     '{\n  "category": "string (필수, 시설|예약|회원권|기타)",\n  "title": "string (필수)",\n  "content": "string (필수)",\n  "imageUrl": "string|null"\n}',
     '{\n  "category": "회원권",\n  "title": "회원권 기간 연장 문의드립니다.",\n  "content": "3개월 정기권 연장 시 할인 혜택이 있나요?",\n  "imageUrl": null\n}',
     '{\n  "inquiryId": "long",\n  "status": "RECEIVED",\n  "createdAt": "datetime"\n}',
     '{\n  "inquiryId": 5,\n  "status": "RECEIVED",\n  "createdAt": "2026-02-17T14:30:00"\n}',
     "201 Created\n400 Bad Request",
     "제목 누락 → 400\n내용 누락 → 400\n카테고리 미선택 → 400",
     "접수 완료 토스트 표시"],

    # RTN-API-003: 루틴 상세
    ["RTN-API-003", "GET", "/api/v1/routines/{routineId}", "루틴 상세 조회",
     "routineId: long (필수)", "-",
     "-",
     "GET /api/v1/routines/1",
     '{\n  "routineId": "long",\n  "name": "string",\n  "goal": "MUSCLE_GAIN|DIET|STAMINA",\n  "difficulty": "BEGINNER|INTERMEDIATE|ADVANCED",\n  "estimatedMin": "int",\n  "frequency": "string",\n  "description": "string",\n  "imageUrl": "string|null",\n  "isRecommended": "boolean",\n  "isFavorited": "boolean",\n  "exercises": [\n    {\n      "exerciseId": "long",\n      "name": "string",\n      "bodyPart": "string",\n      "equipment": "string",\n      "imageUrl": "string|null",\n      "targetSets": "int",\n      "targetRepsMin": "int",\n      "targetRepsMax": "int|null",\n      "orderIndex": "int"\n    }\n  ]\n}',
     '{\n  "routineId": 1,\n  "name": "초보자 파워 빌딩",\n  "goal": "MUSCLE_GAIN",\n  "difficulty": "BEGINNER",\n  "estimatedMin": 45,\n  "frequency": "주 3회",\n  "description": "기초 근력 향상과...",\n  "imageUrl": "/images/routine1.jpg",\n  "isRecommended": true,\n  "isFavorited": false,\n  "exercises": [\n    {\n      "exerciseId": 1,\n      "name": "벤치 프레스",\n      "bodyPart": "가슴",\n      "equipment": "바벨",\n      "imageUrl": null,\n      "targetSets": 3,\n      "targetRepsMin": 10,\n      "targetRepsMax": 12,\n      "orderIndex": 1\n    }\n  ]\n}',
     "200 OK\n404 Not Found",
     "존재하지 않는 루틴 → 404",
     "isFavorited: 로그인 사용자 기준 찜 여부"],

    # USER-API-001: 내 정보 조회
    ["USER-API-001", "GET", "/api/v1/users/me", "내 정보 조회",
     "-", "-",
     "-",
     "GET /api/v1/users/me\nAuthorization: Bearer eyJhbG...",
     '{\n  "userId": "long",\n  "name": "string",\n  "email": "string",\n  "phone": "string",\n  "profileImageUrl": "string|null",\n  "membership": {\n    "type": "string",\n    "dDay": "int",\n    "status": "string"\n  },\n  "attendanceRate": "double (%)",\n  "todayVolume": "double (kg)"\n}',
     '{\n  "userId": 1,\n  "name": "김스미스",\n  "email": "smith@example.com",\n  "phone": "010-1234-5678",\n  "profileImageUrl": null,\n  "membership": {\n    "type": "3개월 정기권",\n    "dDay": 42,\n    "status": "ACTIVE"\n  },\n  "attendanceRate": 92.0,\n  "todayVolume": 1124.0\n}',
     "200 OK\n401 Unauthorized",
     "미인증 요청 → 401",
     "내정보 화면 통합 데이터"],

    # SETT-API-002: 설정 변경
    ["SETT-API-002", "PATCH", "/api/v1/settings", "설정 변경",
     "-", "-",
     '{\n  "notificationEnabled": "boolean (선택)",\n  "darkMode": "SYSTEM|ON|OFF (선택)",\n  "language": "ko|en|ja (선택)"\n}',
     '{\n  "darkMode": "ON"\n}',
     '{\n  "notificationEnabled": "boolean",\n  "darkMode": "string",\n  "language": "string",\n  "updatedAt": "datetime"\n}',
     '{\n  "notificationEnabled": true,\n  "darkMode": "ON",\n  "language": "ko",\n  "updatedAt": "2026-02-17T15:00:00"\n}',
     "200 OK\n400 Bad Request",
     "잘못된 darkMode 값 → 400\n잘못된 language 코드 → 400",
     "부분 수정(PATCH) - 전달된 필드만 변경\n즉시 반영"],

    # RPT-API-001: 운동 리포트 조회
    ["RPT-API-001", "GET", "/api/v1/workout-reports/{sessionId}", "운동 리포트 조회",
     "sessionId: long (필수)", "-",
     "-",
     "GET /api/v1/workout-reports/10",
     '{\n  "reportId": "long",\n  "reportDate": "date",\n  "userName": "string",\n  "totalTimeSec": "int",\n  "totalTimeFormatted": "string (MM:SS)",\n  "totalVolumeKg": "double",\n  "totalCalories": "double",\n  "weeklyChangePct": "double|null",\n  "motivationMsg": "string",\n  "exercises": [\n    {\n      "exerciseName": "string",\n      "sets": "int",\n      "reps": "int",\n      "weightKg": "double",\n      "isCompleted": "boolean"\n    }\n  ],\n  "sharedAt": "datetime|null"\n}',
     '{\n  "reportId": 5,\n  "reportDate": "2026-02-17",\n  "userName": "김스미스",\n  "totalTimeSec": 2712,\n  "totalTimeFormatted": "45:12",\n  "totalVolumeKg": 4250.0,\n  "totalCalories": 380.0,\n  "weeklyChangePct": 12.0,\n  "motivationMsg": "오늘도 한계를 넘으셨군요!",\n  "exercises": [\n    {\n      "exerciseName": "백 스쿼트",\n      "sets": 4,\n      "reps": 10,\n      "weightKg": 100.0,\n      "isCompleted": true\n    }\n  ],\n  "sharedAt": null\n}',
     "200 OK\n404 Not Found",
     "세션 미완료 → 404\n다른 사용자 세션 → 403",
     "메달 아이콘 + 동기부여 메시지\n주간 변화율 자동 계산"],
]

style_sheet(ws2, headers2, widths2)
write_rows(ws2, detail_data)


# ============================================================
# Sheet 3: 공통 에러 코드
# ============================================================
ws3 = wb.create_sheet("공통 에러 코드")
headers3 = ["HTTP 상태코드", "에러 코드", "메시지 (ko)", "설명", "대응 방법"]
widths3 = [14, 18, 30, 40, 35]

error_data = [
    ["200 OK", "-", "성공", "요청 정상 처리", "-"],
    ["201 Created", "-", "생성 완료", "리소스 생성 성공 (회원가입, 예약 등)", "-"],
    ["400 Bad Request", "INVALID_INPUT", "입력값이 올바르지 않습니다.", "필수 값 누락, 형식 오류, 유효성 검증 실패", "입력 필드 유효성 재확인"],
    ["400 Bad Request", "INVALID_FORMAT", "형식이 올바르지 않습니다.", "이메일/전화번호 등 형식 오류", "정규식 기반 클라이언트 검증"],
    ["400 Bad Request", "PASSWORD_TOO_SHORT", "비밀번호는 6자 이상이어야 합니다.", "비밀번호 최소 길이 미충족", "6자 이상 재입력 안내"],
    ["401 Unauthorized", "INVALID_CREDENTIALS", "아이디 또는 비밀번호가 올바르지 않습니다.", "로그인 실패 - 이메일/비밀번호 불일치", "재입력 안내"],
    ["401 Unauthorized", "TOKEN_EXPIRED", "인증이 만료되었습니다.", "Access Token 만료", "Refresh Token으로 갱신 시도"],
    ["401 Unauthorized", "TOKEN_INVALID", "유효하지 않은 인증입니다.", "JWT 토큰 위변조/무효", "로그인 화면으로 이동"],
    ["403 Forbidden", "ACCESS_DENIED", "접근 권한이 없습니다.", "타인의 리소스 접근 또는 비활성 회원", "권한 확인"],
    ["403 Forbidden", "MEMBERSHIP_EXPIRED", "회원권이 만료되었습니다.", "회원권 만료 상태에서 시설 이용 시도", "회원권 연장 안내"],
    ["404 Not Found", "RESOURCE_NOT_FOUND", "요청한 정보를 찾을 수 없습니다.", "존재하지 않는 리소스 ID", "ID 확인"],
    ["409 Conflict", "DUPLICATE_EMAIL", "이미 가입된 이메일입니다.", "회원가입 이메일 중복", "로그인 화면 유도"],
    ["409 Conflict", "DUPLICATE_PHONE", "이미 등록된 전화번호입니다.", "회원가입 전화번호 중복", "로그인 화면 유도"],
    ["409 Conflict", "SLOT_FULL", "해당 시간은 만석입니다.", "예약 시 슬롯 최대 인원 초과", "다른 시간 안내"],
    ["409 Conflict", "DUPLICATE_RESERVATION", "이미 예약된 시간입니다.", "동일 슬롯 중복 예약 시도", "기존 예약 확인 안내"],
    ["409 Conflict", "ACTIVE_SESSION_EXISTS", "이미 진행 중인 세션이 있습니다.", "활성 세션 존재 시 새 세션 생성 시도", "기존 세션 종료 후 재시도"],
    ["429 Too Many Requests", "RATE_LIMITED", "요청이 너무 많습니다. 잠시 후 다시 시도해주세요.", "API 호출 빈도 초과", "잠시 대기 후 재시도"],
    ["500 Internal Server Error", "SERVER_ERROR", "서버 오류가 발생했습니다.", "서버 내부 오류", "재시도 + 관리자 문의 안내"],
    ["503 Service Unavailable", "SERVICE_UNAVAILABLE", "서비스 점검 중입니다.", "서버 점검/배포 중", "잠시 후 재시도 안내"],
]

style_sheet(ws3, headers3, widths3)
write_rows(ws3, error_data)


# ============================================================
# Sheet 4: 공통 Request/Response 규격
# ============================================================
ws4 = wb.create_sheet("공통 규격")
headers4 = ["항목", "내용", "예시", "비고"]
widths4 = [22, 50, 45, 30]

common_data = [
    ["Base URL", "https://api.smithlife.co.kr/api/v1", "https://api.smithlife.co.kr/api/v1/auth/login", "개발 환경: http://localhost:8080/api/v1"],
    ["API 버전", "URL Path 방식 (v1)", "/api/v1/...", "버전 업 시 /api/v2/... 신설"],
    ["인증 방식", "Bearer Token (JWT)", 'Authorization: Bearer eyJhbGciOi...', "Access Token: 1시간, Refresh Token: 14일"],
    ["Content-Type", "application/json (기본)\nmultipart/form-data (파일 업로드)", 'Content-Type: application/json; charset=utf-8', "이미지 업로드 API만 multipart 사용"],
    ["Accept-Language", "다국어 지원 (ko, en, ja)", "Accept-Language: ko", "에러 메시지 다국어 대응"],
    ["날짜 형식", "ISO 8601", "2026-02-17T14:30:00", "Date: yyyy-MM-dd, DateTime: yyyy-MM-dd'T'HH:mm:ss"],
    ["시간 형식", "HH:mm (24시간제)", "09:00, 14:30", "예약 시간표/슬롯에 사용"],
    ["페이지네이션", "Offset 기반 (page/size)", "?page=0&size=20", "page: 0부터 시작, size: 기본 20"],
    ["페이지 응답 구조", '{"content":[], "totalElements":n, "totalPages":n, "hasNext":bool}', '페이지 응답 예시 참조', "Spring Page 구조 호환"],
    ["정렬", "기본 최신순 (createdAt DESC)", "?sort=createdAt,desc", "필요 시 sort 파라미터 지원"],
    ["에러 응답 구조", '{"code":"string", "message":"string", "details":{}|null}', '{"code":"INVALID_INPUT","message":"이메일 형식이 올바르지 않습니다.","details":{"field":"email"}}', "모든 4xx/5xx 에러에 동일 구조 적용"],
    ["빈 배열 응답", "데이터 없을 시 빈 배열 반환 (null 아님)", '{"content": [], "totalElements": 0}', "클라이언트 빈 상태(Empty State) UI 판단용"],
    ["Boolean 응답", "true/false (소문자)", '{"available": true}', "JSON boolean 표준"],
    ["CORS", "허용 Origin 화이트리스트", "Access-Control-Allow-Origin: https://smithlife.co.kr", "개발 환경: localhost 허용"],
    ["Rate Limiting", "IP/User 당 분당 60회 (기본)", "X-RateLimit-Remaining: 55", "비밀번호 재설정: 분당 3회 제한"],
    ["HTTPS", "TLS 1.2 이상 필수", "https://...", "NFR-006 준수"],
]

style_sheet(ws4, headers4, widths4)
write_rows(ws4, common_data)


# ============================================================
# Sheet 5: DB 테이블-API 매핑
# ============================================================
ws5 = wb.create_sheet("DB-API 매핑")
headers5 = ["No", "DB 테이블", "테이블 설명", "관련 API ID 목록", "주요 CRUD", "비고"]
widths5 = [5, 20, 25, 45, 15, 25]

mapping_data = [
    [1, "user", "회원", "AUTH-API-001~008, USER-API-001~005", "CRUD", "회원가입/로그인/프로필 수정/탈퇴"],
    [2, "membership", "회원권", "MBSP-API-001~003, HOME-API-001", "CRU", "회원권 등록/조회/일시정지"],
    [3, "user_settings", "사용자 설정", "SETT-API-001~002", "RU", "설정 조회/변경"],
    [4, "attendance", "출석 기록", "ATT-API-001~005", "CRU", "QR체크인/체크아웃/출석조회"],
    [5, "facility", "시설", "FAC-API-001~002, SLOT-API-001", "R", "시설 목록/혼잡도 조회"],
    [6, "time_slot", "시간 슬롯", "SLOT-API-001, RESV-API-001", "RU", "시간표 조회/예약 시 count 변경"],
    [7, "reservation", "예약", "RESV-API-001~005, HOME-API-001", "CRU", "예약 생성/조회/취소"],
    [8, "usage_history", "이용내역", "USAGE-API-001~002", "CR", "이용 기록 자동 생성/조회"],
    [9, "exercise", "운동 종목", "EXER-API-001~002", "R", "운동 검색/상세 조회"],
    [10, "routine", "루틴", "RTN-API-001~007", "CRUD", "루틴 CRUD + 추천 루틴"],
    [11, "routine_exercise", "루틴-운동 구성", "RTN-API-003~005", "CRD", "루틴 운동 구성 관리"],
    [12, "favorite_routine", "루틴 찜", "FAV-API-001~003", "CRD", "찜 추가/해제/목록"],
    [13, "workout_session", "운동 세션", "SESS-API-001~005", "CRU", "세션 시작/종료/일시정지"],
    [14, "session_exercise", "세션-운동", "SE-API-001~005", "CRUD", "세션 내 운동 관리"],
    [15, "exercise_set", "세트 기록", "SET-API-001~005", "CRUD", "세트 추가/수정/완료/삭제"],
    [16, "personal_record", "개인 최고 기록", "PR-API-001~002", "CR", "세트 완료 시 자동 갱신/조회"],
    [17, "workout_report", "운동 리포트", "RPT-API-001~004", "CR", "세션 종료 시 자동 생성/조회"],
    [18, "notification", "알림", "NOTI-API-001~004", "CRU", "알림 생성/조회/읽음처리"],
    [19, "announcement", "공지사항", "ANN-API-001~002, HOME-API-002", "R", "공지 목록/상세 조회"],
    [20, "inquiry", "1:1 문의", "INQ-API-001~004", "CR", "문의 작성/목록/상세 조회"],
    [21, "inquiry_reply", "문의 답변", "INQ-API-003", "R", "문의 상세 시 답변 함께 조회"],
    [22, "faq", "자주 묻는 질문", "FAQ-API-001", "R", "FAQ 목록 조회"],
]

style_sheet(ws5, headers5, widths5)
write_rows(ws5, mapping_data)


# ── 저장 ──
output_path = r"c:\SSAFY\personalProject\SmithLife2\SmithLife_API명세서_v1.0.xlsx"
wb.save(output_path)
print(f"API 명세서 생성 완료: {output_path}")
print(f"  - Sheet 1: API 총괄 목록 ({len(api_list)}개 API)")
print(f"  - Sheet 2: API 상세 명세 ({len(detail_data)}개 주요 API 상세)")
print(f"  - Sheet 3: 공통 에러 코드 ({len(error_data)}개)")
print(f"  - Sheet 4: 공통 규격 ({len(common_data)}개 항목)")
print(f"  - Sheet 5: DB-API 매핑 ({len(mapping_data)}개 테이블)")
